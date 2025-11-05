from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q, Count, F
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from decimal import Decimal, InvalidOperation
import json
import openpyxl
from io import BytesIO
from django.db import connection
from django.utils import timezone

from .models import Member, Transaction, OrganizationUser, Organization, OrganizationTheme, MemberEditLog, PaymentRequest, SystemSettings
from .forms import CustomLoginForm, MemberForm, QuickMemberForm, TransactionForm, MemberUpdateForm, ExcelImportForm, SignUpForm, AddOrganizationUserForm
# Removed Django's staff_member_required - using org_staff_required instead
from django.contrib.auth.models import User
from .permissions import org_staff_required, org_admin_required, org_owner_required, is_org_owner, is_org_admin
import secrets
import string

# ============================================================================
# PUBLIC VIEWS (Landing, Signup)
# ============================================================================

def landing_view(request):
    """Landing page for new visitors"""
    if request.user.is_authenticated:
        return redirect_to_dashboard(request)
    return render(request, 'landing.html')


def help_center(request, org_slug):
    """Tenant Help Center - FAQs and guides for the current organization (non-technical)."""
    organization = get_object_or_404(Organization, slug=org_slug, is_active=True)
    faqs = [
        {
            'q': 'How do I add members?',
            'a': 'If you have permission, go to Admin > Staff Management. Owners may also use Import (Excel) if enabled.'
        },
        {
            'q': 'How do I record daily collections?',
            'a': 'Open Daily Collection from the navbar, search a member, enter amount, and save.'
        },
        {
            'q': 'How do I view reports?',
            'a': 'Use Admin Log to see transactions. Export to Excel/PDF from the Admin menu if allowed.'
        },
        {
            'q': 'I cannot access Admin features',
            'a': 'Contact your organization owner to grant you the required role.'
        },
    ]

    # Support contacts from SystemSettings with fallbacks
    settings_obj = SystemSettings.get_settings() if hasattr(SystemSettings, 'get_settings') else None
    support_email = (settings_obj.support_email if settings_obj and settings_obj.support_email else 'KodinSoftwares@gmail.com')
    support_phone = (settings_obj.support_phone if settings_obj and settings_obj.support_phone else '0614021404')

    # Build tel and WhatsApp links (digits only for wa.me)
    import re
    phone_digits = re.sub(r'\D+', '', support_phone or '')
    whatsapp_link = f"https://wa.me/{phone_digits}" if phone_digits else None

    return render(request, 'tracker/help_center.html', {
        'organization': organization,
        'faqs': faqs,
        'support_email': support_email,
        'support_phone': support_phone,
        'whatsapp_link': whatsapp_link,
    })

def signup_view(request):
    """User registration and organization creation"""
    if request.user.is_authenticated:
        return redirect_to_dashboard(request)
    
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            # Enforce Terms & Conditions acceptance
            if not request.POST.get('accept_terms'):
                messages.error(request, 'You must accept the Terms & Conditions to create an account.')
                return render(request, 'tracker/signup.html', {'form': form})
            try:
                # Create user
                user = User.objects.create_user(
                    username=form.cleaned_data['username'],
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name']
                )
                
                # Create organization
                org_slug = form.cleaned_data['organization_name'].lower().replace(' ', '-')
                # Ensure unique slug
                base_slug = org_slug
                counter = 1
                while Organization.objects.filter(slug=org_slug).exists():
                    org_slug = f"{base_slug}-{counter}"
                    counter += 1
                
                organization = Organization.objects.create(
                    name=form.cleaned_data['organization_name'],
                    slug=org_slug,
                    description=form.cleaned_data['organization_description'],
                    is_active=True
                )
                # Start default free trial status
                organization.trial_started_at = timezone.now()
                organization.subscription_status = 'FREE_TRIAL'
                organization.save(update_fields=['trial_started_at', 'subscription_status'])
                
                # Create organization theme with defaults
                OrganizationTheme.objects.create(
                    organization=organization,
                    primary_color='#7492b9',
                    secondary_color='#6c757d',
                    success_color='#28a745',
                    warning_color='#ffc107',
                    danger_color='#dc3545',
                    navbar_title=organization.name,
                    watermark_text='Bossin',
                    default_pledge_amount=Decimal('70000.00'),
                    target_amount=Decimal('210000.00')
                )
                
                # Create organization user (owner)
                OrganizationUser.objects.create(
                    user=user,
                    organization=organization,
                    role='owner',
                    is_active=True
                )
                
                # Log the user in (handle multiple auth backends)
                auth_user = authenticate(request, username=form.cleaned_data['username'], password=form.cleaned_data['password'])
                if auth_user is not None:
                    login(request, auth_user)
                else:
                    from django.conf import settings
                    # Fallback: explicitly specify a backend
                    login(request, user, backend=settings.AUTHENTICATION_BACKENDS[0])
                messages.success(request, f'Welcome {user.first_name}! Your organization "{organization.name}" has been created.')
                
                # Redirect to onboarding flow
                return redirect('tracker:onboarding_financial', org_slug=organization.slug)
                
            except Exception as e:
                messages.error(request, f'Error creating account: {str(e)}')
    else:
        form = SignUpForm()
    
    return render(request, 'tracker/signup.html', {'form': form})


def terms_view(request):
    """Public Terms & Conditions page."""
    return render(request, 'terms.html')


# Helper function to redirect to dashboard with org_slug
def redirect_to_dashboard(request):
    """Redirect to user's organization dashboard or admin if no org"""
    if hasattr(request, 'tenant') and request.tenant:
        return redirect('tracker:dashboard', org_slug=request.tenant.slug)
    
    # Try to get user's first organization
    org_user = OrganizationUser.objects.filter(user=request.user, is_active=True).first()
    if org_user:
        return redirect('tracker:dashboard', org_slug=org_user.organization.slug)
    
    # Fall back to admin
    return redirect('admin:index')

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from django.db.models import Sum
# Add these imports to your views.py file
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from .models import Member, Transaction
from reportlab.platypus import PageBreak, KeepTogether
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.utils import timezone
from decimal import Decimal
import json

from .models import Member, Transaction
from .forms import MemberForm, TransactionForm


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Q
from decimal import Decimal
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Q, F
from django.db import models
from decimal import Decimal
import json

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, F
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from .models import Member

@org_staff_required
def edit_members(request, org_slug=None):
    """Main view for editing member details with smart filtering"""
    try:
        # Get tenant from request for data isolation
        tenant = request.tenant

        # Get members for THIS ORGANIZATION ONLY
        members = Member.objects.filter(organization=tenant).order_by('name').select_related()

        # Handle search (keep for backend if needed)
        search_query = request.GET.get('search', '').strip()
        if search_query:
            members = members.filter(
                Q(name__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(course__icontains=search_query)
            )

        # Calculate statistics for quick stats - FOR THIS ORGANIZATION ONLY
        total_members = Member.objects.filter(organization=tenant).count()

        # Calculate payment statistics using model properties - FOR THIS ORGANIZATION ONLY
        all_members = Member.objects.filter(organization=tenant)
        completed_payments = 0
        incomplete_payments = 0
        not_started_payments = 0
        exceeded_payments = 0

        for member in all_members:
            if member.is_complete:
                completed_payments += 1
            elif member.is_incomplete:
                incomplete_payments += 1
            elif member.not_started:
                not_started_payments += 1
            elif member.has_exceeded:
                exceeded_payments += 1

        # Prepare members data with all fields and status information
        members_data = []
        for member in members:
            # Determine status for filtering
            if member.is_complete:
                status = 'complete'
                status_html = '<small class="status-complete">✓ Complete</small>'
            elif member.is_incomplete:
                status = 'incomplete'
                status_html = '<small class="status-incomplete">⚠ Incomplete</small>'
            elif member.not_started:
                status = 'not_started'
                status_html = '<small class="status-not-started">✗ Not Started</small>'
            elif member.has_exceeded:
                status = 'exceeded'
                status_html = '<small class="status-exceeded">↗ Exceeded</small>'
            else:
                status = 'unknown'
                status_html = '<small class="text-muted">-</small>'

            members_data.append({
                'id': member.id,
                'name': member.name,
                'phone': member.phone or '',
                'email': member.email or '',
                'course': member.course or '',
                'year_of_study': member.year or '',
                'pledge': member.pledge,
                'paid_total': member.paid_total,
                'remaining': member.remaining,
                'is_active': member.is_active,
                'status_display': member.status_display,
                'is_complete': member.is_complete,
                'is_incomplete': member.is_incomplete,
                'not_started': member.not_started,
                'has_exceeded': member.has_exceeded,
                'status': status,
                'status_html': status_html,
                'created_at': member.created_at,
                'updated_at': member.updated_at,
            })

        context = {
            'members': members_data,
            'search_query': search_query,
            'total_members': total_members,
            'completed_payments': completed_payments,
            'incomplete_payments': incomplete_payments,
            'not_started_payments': not_started_payments,
            'exceeded_payments': exceeded_payments,
            'can_edit': True,  # Admin can always edit (decorator ensures this)
        }

        return render(request, 'tracker/Edit_Members_table.html', context)

    except Exception as e:
        messages.error(request, f'Error loading members: {str(e)}')
        return redirect_to_dashboard(request)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def update_member_ajaxs(request, org_slug=None):
    """AJAX endpoint for updating member details"""
    try:
        # Get the tenant from request
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return JsonResponse({'success': False, 'error': 'Organization not found'})

        data = json.loads(request.body)
        member_id = data.get('member_id')

        if not member_id:
            return JsonResponse({'success': False, 'error': 'Member ID is required'})

        try:
            # Filter by organization for data isolation
            member = Member.objects.get(id=member_id, organization=tenant)
        except Member.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Member not found'})

        # Track changes for logging
        changes = []
        
        # Update basic fields (always allowed)
        if 'name' in data:
            if not data['name'].strip():
                return JsonResponse({'success': False, 'error': 'Name cannot be empty'})
            new_name = data['name'].strip()
            if member.name != new_name:
                changes.append(('name', member.name, new_name))
                member.name = new_name

        if 'phone_number' in data:
            new_phone = data['phone_number'].strip()
            if member.phone != new_phone:
                changes.append(('phone', member.phone, new_phone))
                member.phone = new_phone

        if 'year_of_study' in data:
            new_year = data['year_of_study'].strip()
            if member.year != new_year:
                changes.append(('year', member.year, new_year))
                member.year = new_year

        # Update paid amount (admin and owner only)
        if 'paid_total' in data:
            if not is_org_admin(request.user, tenant):
                # If non-admin user somehow sent paid_total, ignore it and continue
                # This prevents errors while maintaining security
                data.pop('paid_total', None)
            else:
                try:
                    paid_amount = float(data['paid_total'])
                    if paid_amount < 0:
                        return JsonResponse({'success': False, 'error': 'Paid amount cannot be negative'})
                    if member.paid_total != paid_amount:
                        changes.append(('paid_total', str(member.paid_total), str(paid_amount)))
                        member.paid_total = paid_amount
                except (ValueError, TypeError):
                    return JsonResponse({'success': False, 'error': 'Invalid paid amount'})

        member.save()
        
        # Log all changes to MemberEditLog
        for field_name, before_val, after_val in changes:
            MemberEditLog.objects.create(
                organization=tenant,
                member=member,
                field_changed=field_name,
                before_value=before_val,
                after_value=after_val,
                edited_by=request.user
            )

        # Calculate updated status for response
        if member.is_complete:
            status = 'complete'
            status_html = '<small class="status-complete">✓ Complete</small>'
        elif member.is_incomplete:
            status = 'incomplete'
            status_html = '<small class="status-incomplete">⚠ Incomplete</small>'
        elif member.not_started:
            status = 'not_started'
            status_html = '<small class="status-not-started">✗ Not Started</small>'
        elif member.has_exceeded:
            status = 'exceeded'
            status_html = '<small class="status-exceeded">↗ Exceeded</small>'
        else:
            status = 'unknown'
            status_html = '<small class="text-muted">-</small>'

        return JsonResponse({
            'success': True,
            'message': f'{member.name} updated successfully',
            'member': {
                'id': member.id,
                'name': member.name,
                'paid_total': member.paid_total,
                'status': status,
                'status_html': status_html,
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def add_member_ajaxs(request, org_slug=None):
    """AJAX endpoint for adding new members"""
    try:
        # Get the tenant from request
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return JsonResponse({'success': False, 'error': 'Organization not found'})

        data = json.loads(request.body)

        # Validate required fields
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})

        # Check if member already exists in THIS organization only
        if Member.objects.filter(organization=tenant, name__iexact=name).exists():
            return JsonResponse({'success': False, 'error': 'Member with this name already exists'})

        # Get pledge amount (required)
        pledge = data.get('pledge', 70000)
        try:
            pledge = float(pledge)
            if pledge < 0:
                return JsonResponse({'success': False, 'error': 'Pledge amount cannot be negative'})
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid pledge amount'})

        # Create new member with organization
        member = Member(
            organization=tenant,
            name=name,
            pledge=pledge,
            phone=data.get('phone_number', '').strip(),
        )

        member.save()

        return JsonResponse({
            'success': True,
            'message': f'{member.name} added successfully',
            'member': {
                'id': member.id,
                'name': member.name,
                'phone': member.phone,
                'pledge': member.pledge,
                'paid_total': member.paid_total,
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def delete_member_ajaxs(request, org_slug=None):
    """AJAX endpoint for deleting members (admin only)"""
    # Get the tenant from request first
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Organization not found'})
    
    # Check if user is admin or owner
    if not is_org_admin(request.user, tenant):
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    try:
        data = json.loads(request.body)
        member_id = data.get('member_id')

        if not member_id:
            return JsonResponse({'success': False, 'error': 'Member ID is required'})

        try:
            # Filter by organization for data isolation
            member = Member.objects.get(id=member_id, organization=tenant)
            member_name = member.name
            member.delete()

            return JsonResponse({
                'success': True,
                'message': f'{member_name} deleted successfully'
            })

        except Member.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Member not found'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})


# Helper function to validate and format phone numbers
def format_phone_number(phone_number):
    """Format phone number to international format"""
    if not phone_number:
        return None

    phone_number = phone_number.strip()

    # Remove spaces and hyphens
    phone_number = phone_number.replace(' ', '').replace('-', '')

    # Format to Tanzania international format
    if not phone_number.startswith('+'):
        if phone_number.startswith('0'):
            phone_number = '+255' + phone_number[1:]
        elif phone_number.startswith('255'):
            phone_number = '+' + phone_number
        elif phone_number.isdigit() and len(phone_number) >= 9:
            phone_number = '+255' + phone_number

    return phone_number


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def toggle_member_status_ajax(request):
    """AJAX view to toggle member active status"""
    try:
        data = json.loads(request.body)
        member_id = data.get('member_id')

        if not member_id:
            return JsonResponse({
                'success': False,
                'error': 'Member ID is required'
            })

        member = get_object_or_404(Member, id=member_id)

        # Toggle active status
        member.is_active = not member.is_active
        member.save()

        status = "activated" if member.is_active else "deactivated"

        return JsonResponse({
            'success': True,
            'message': f'Member {member.name} has been {status}',
            'is_active': member.is_active
        })

    except Member.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Member not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        })

@login_required
def export_excel(request, org_slug=None):
    """Export all members data to Excel with styling - ORGANIZATION DATA ONLY"""
    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            messages.error(request, 'Organization not found')
            return redirect_to_dashboard(request)
        
        # Check if user is member of this organization
        if not OrganizationUser.objects.filter(user=request.user, organization=tenant, is_active=True).exists():
            messages.error(request, 'You are not a member of this organization')
            return redirect_to_dashboard(request)
        
        # Get all members with the same filtering logic as dashboard
        search_query = request.GET.get('search', '')
        filter_status = request.GET.get('filter', '')

        # Get members FOR THIS ORGANIZATION ONLY
        members = []
        try:
            all_members = Member.objects.filter(organization=tenant)
            for member in all_members:
                try:
                    if member.pledge is None or member.paid_total is None:
                        if member.pledge is None:
                            member.pledge = Decimal('70000.00')
                        if member.paid_total is None:
                            member.paid_total = Decimal('0.00')
                        member.save()
                    members.append(member)
                except Exception:
                    continue
        except Exception:
            members = []

        # Apply search filter
        if search_query and members:
            filtered_members = []
            for member in members:
                try:
                    if (search_query.lower() in member.name.lower() or
                        (member.phone and search_query in member.phone) or
                        (member.email and search_query.lower() in member.email.lower())):
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        # Apply status filter
        if filter_status and members:
            filtered_members = []
            for member in members:
                try:
                    if filter_status == 'incomplete' and member.paid_total > 0 and member.paid_total < member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'complete' and member.paid_total >= member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'pledged' and member.pledge > 70000:
                        filtered_members.append(member)
                    elif filter_status == 'not_started' and member.paid_total == 0:
                        filtered_members.append(member)
                    elif filter_status == 'exceeded' and member.paid_total > member.pledge:
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members Data"

        # Add organization header
        ws.merge_cells('A1:J1')
        org_header = ws['A1']
        org_header.value = f"{tenant.name} - Members Report"
        org_header.font = Font(bold=True, size=14, color="FFFFFF")
        org_header.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        org_header.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Add export date
        ws.merge_cells('A2:J2')
        date_cell = ws['A2']
        date_cell.value = f"Exported on {date.today().strftime('%B %d, %Y')}"
        date_cell.font = Font(italic=True, size=10, color="666666")
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 15

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        currency_alignment = Alignment(horizontal='right', vertical='center')

        # Headers (starting at row 4 after org header and date)
        headers = [
            'Name', 'Phone', 'Email', 'Course', 'Year',
            'Pledge Amount', 'Paid Amount', 'Balance', 'Status', 'Progress %'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Data rows (starting at row 5)
        for row_num, member in enumerate(members, 5):
            try:
                pledge = member.pledge if member.pledge is not None else Decimal('70000.00')
                paid = member.paid_total if member.paid_total is not None else Decimal('0.00')
                balance = pledge - paid
                progress = (paid / pledge * 100) if pledge > 0 else 0

                # Determine status
                if paid == 0:
                    status = "Not Started"
                elif paid < pledge:
                    status = "Incomplete"
                elif paid == pledge:
                    status = "Complete"
                else:
                    status = "Exceeded"

                # Data
                data = [
                    member.name,
                    member.phone or '',
                    member.email or '',
                    member.course or '',
                    member.year or '',
                    float(pledge),
                    float(paid),
                    float(balance),
                    status,
                    f"{progress:.1f}%"
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border

                    # Special formatting for currency columns
                    if col in [6, 7, 8]:  # Pledge, Paid, Balance
                        cell.number_format = '#,##0.00'
                        cell.alignment = currency_alignment
                    else:
                        cell.alignment = center_alignment

            except Exception:
                continue

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")

        # Calculate totals
        total_members = len(members)
        total_pledged = sum(member.pledge or Decimal('70000.00') for member in members)
        total_collected = sum(member.paid_total or Decimal('0.00') for member in members)
        # Get organization's target amount from theme
        try:
            target_amount = Decimal(str(tenant.theme.target_amount))
        except:
            target_amount = Decimal('210000.00')
        progress_percentage = (total_collected / target_amount * 100) if target_amount > 0 else 0

        # Summary data
        summary_data = [
            ['FUNDRAISING SUMMARY', ''],
            ['', ''],
            ['Total Members', total_members],
            ['Total Pledged', f"TSh {total_pledged:,.2f}"],
            ['Total Collected', f"TSh {total_collected:,.2f}"],
            ['Target Amount', f"TSh {target_amount:,.2f}"],
            ['Progress', f"{progress_percentage:.1f}%"],
            ['', ''],
            ['Export Date', date.today().strftime('%Y-%m-%d')],
        ]

        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_num, column=2, value=value)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Include organization name in filename
        org_name_slug = tenant.slug.replace('-', '_')
        filename = f"{org_name_slug}_members_data_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        messages.error(request, f"Error exporting Excel: {str(e)}")
        return redirect_to_dashboard(request)
@login_required
def export_pdf(request, org_slug=None):
    """Export all members data to PDF with styling, logo, and dynamic headings - ORGANIZATION DATA ONLY"""
    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            messages.error(request, 'Organization not found')
            return redirect_to_dashboard(request)
        
        # Check if user is member of this organization
        if not OrganizationUser.objects.filter(user=request.user, organization=tenant, is_active=True).exists():
            messages.error(request, 'You are not a member of this organization')
            return redirect_to_dashboard(request)
        
        # Get all members with the same filtering logic as dashboard
        search_query = request.GET.get('search', '')
        filter_status = request.GET.get('filter', '')

        # Get members FOR THIS ORGANIZATION ONLY
        members = []
        try:
            all_members = Member.objects.filter(organization=tenant)
            for member in all_members:
                try:
                    if member.pledge is None or member.paid_total is None:
                        if member.pledge is None:
                            member.pledge = Decimal('70000.00')
                        if member.paid_total is None:
                            member.paid_total = Decimal('0.00')
                        member.save()
                    members.append(member)
                except Exception:
                    continue
        except Exception:
            members = []

        # Apply same filters as Excel export
        if search_query and members:
            filtered_members = []
            for member in members:
                try:
                    if (search_query.lower() in member.name.lower() or
                        (member.phone and search_query in member.phone) or
                        (member.email and search_query.lower() in member.email.lower())):
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        if filter_status and members:
            filtered_members = []
            for member in members:
                try:
                    if filter_status == 'incomplete' and member.paid_total > 0 and member.paid_total < member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'complete' and member.paid_total >= member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'pledged' and member.pledge > 70000:
                        filtered_members.append(member)
                    elif filter_status == 'not_started' and member.paid_total == 0:
                        filtered_members.append(member)
                    elif filter_status == 'exceeded' and member.paid_total > member.pledge:
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        # Create PDF with custom page template for logo
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=120, bottomMargin=60)

        # Container for PDF elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#2c3e50')
        )

        filter_style = ParagraphStyle(
            'FilterStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#7f8c8d')
        )

        # Link style for download hyperlink
        link_style = ParagraphStyle(
            'LinkStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=15,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#3498db')
        )

        # Dynamic title based on filters - Use organization name
        base_title = f"MEMBERS REPORT: {tenant.name.upper()}"

        # Add filter information to title
        filter_info = []
        if filter_status:
            filter_map = {
                'incomplete': 'INCOMPLETE PAYMENTS',
                'complete': 'COMPLETE PAYMENTS',
                'pledged': 'PLEDGED ABOVE Tsh 70,000',
                'not_started': 'NOT STARTED',
                'exceeded': 'EXCEEDED PLEDGES'
            }
            filter_info.append(filter_map.get(filter_status, filter_status.upper()))

        if search_query:
            filter_info.append(f'SEARCH: "{search_query.upper()}"')

        if filter_info:
            title_text = f"{base_title}<br/><font size='14' color='#e74c3c'>({' - '.join(filter_info)})</font>"
        else:
            title_text = base_title

        # Title
        title = Paragraph(title_text, title_style)
        elements.append(title)

        # Add download link for latest version
        current_url = request.build_absolute_uri()
        download_link_text = f'<a href="{current_url}" color="#3498db"><u>Click here to download the latest version of this report</u></a>'
        download_link = Paragraph(download_link_text, link_style)
        elements.append(download_link)
        elements.append(Spacer(1, 12))

        # Summary section
        total_members = len(members)
        total_pledged = sum(member.pledge or Decimal('70000.00') for member in members)
        total_collected = sum(member.paid_total or Decimal('0.00') for member in members)
        # Get organization's target amount from theme
        try:
            target_amount = Decimal(str(tenant.theme.target_amount))
        except:
            target_amount = Decimal('210000.00')
        progress_percentage = (total_collected / target_amount * 100) if target_amount > 0 else 0

        summary_data = [
            ['SUMMARY', ''],
            ['Total Members:', f"{total_members:,}"],
            ['Total Pledged:', f"TSh {total_pledged:,.2f}"],
            ['Total Collected:', f"TSh {total_collected:,.2f}"],
            ['Target Amount:', f"TSh {target_amount:,.2f}"],
            ['% Collected:', f"{progress_percentage:.1f}%"],
            ['Report Date:', date.today().strftime('%B %d, %Y')],
        ]

        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Members table
        if members:
            # Table headers with index number
            headers = ['#', 'Name', 'Phone', 'Pledge', 'Paid', 'Exceed/Remain(-)', 'Status']

            # Prepare data
            table_data = [headers]

            for index, member in enumerate(members, 1):
                try:
                    pledge = member.pledge if member.pledge is not None else Decimal('70000.00')
                    paid = member.paid_total if member.paid_total is not None else Decimal('0.00')
                    balance = paid - pledge

                    if paid == 0:
                        status = "Not Started"
                    elif paid < pledge:
                        status = "Incomplete"
                    elif paid == pledge:
                        status = "Complete"
                    else:
                        status = "Exceeded"

                    row = [
                        str(index),  # Index number
                        member.name[:20] + "..." if len(member.name) > 20 else member.name,
                        member.phone[:15] if member.phone else '',
                        f"TSh {pledge:,.0f}",
                        f"TSh {paid:,.0f}",
                        f"TSh {balance:,.0f}",
                        status
                    ]
                    table_data.append(row)
                except Exception:
                    continue

            # Create table with adjusted column widths to accommodate index
            table = Table(table_data, colWidths=[0.4*inch, 1.3*inch, 1*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),

                # Data styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),

                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),

                # Currency column alignment
                ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
                # Index column alignment
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ]))

            # With this:
            if len(table_data) > 1:  # Only if there's data beyond headers
                # For very large tables, you might want to split them
                if len(table_data) > 36:  # Adjust this number based on your needs
                    # Split large tables into chunks
                    chunk_size = 200
                    for i in range(0, len(table_data), chunk_size):
                        if i == 0:
                            chunk_data = table_data[0:chunk_size+1]  # Include header
                        else:
                            chunk_data = [table_data[0]] + table_data[i:i+chunk_size]  # Header + data

                        chunk_table = Table(chunk_data, colWidths=[0.4*inch, 1.3*inch, 1*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
                        chunk_table.setStyle(TableStyle([
                            # Header styling
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                            ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
                            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                        ]))

                        elements.append(chunk_table)
                        if i + chunk_size < len(table_data) - 1:  # Not the last chunk
                            elements.append(PageBreak())
                else:
                    # For smaller tables, just add normally
                    elements.append(table)

# Custom page template with organization logo and watermark
        def add_logo_and_header(canvas, doc):
            """Add organization logo, header, and watermark to each page"""
            from django.conf import settings
            import os
            from reportlab.lib.utils import ImageReader
            from PIL import Image
            
            canvas.saveState()

            # Set white background for entire page
            canvas.setFillColor(colors.white)
            canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)

            # Try to add organization logo first
            logo_path = None
            if tenant.theme and tenant.theme.logo:
                try:
                    logo_path = tenant.theme.logo.path
                except:
                    logo_path = None
            
            if logo_path and os.path.exists(logo_path):
                try:
                    # Open image with PIL to handle transparency
                    pil_image = Image.open(logo_path)

                    # Convert to RGB if it has transparency (RGBA) or is in other modes
                    if pil_image.mode in ('RGBA', 'LA', 'P'):
                        # Create white background
                        background = Image.new('RGB', pil_image.size, (255, 255, 255))
                        if pil_image.mode == 'P':
                            pil_image = pil_image.convert('RGBA')
                        background.paste(pil_image, mask=pil_image.split()[-1] if pil_image.mode in ('RGBA', 'LA') else None)
                        pil_image = background
                    elif pil_image.mode != 'RGB':
                        pil_image = pil_image.convert('RGB')

                    logo = ImageReader(pil_image)
                    # Get original dimensions
                    logo_width, logo_height = pil_image.size

                    # Calculate scaled dimensions (max 100x50)
                    max_width, max_height = 100, 50
                    if logo_width > max_width or logo_height > max_height:
                        ratio = min(max_width/logo_width, max_height/logo_height)
                        logo_width *= ratio
                        logo_height *= ratio

                    # Position logo at top center
                    x = (A4[0] - logo_width) / 2  # Center horizontally
                    y = A4[1] - 80  # 80 points from top

                    # Draw the organization logo
                    canvas.drawImage(logo, x, y, width=logo_width, height=logo_height)
                except Exception as e:
                    # If logo fails, add text header instead
                    canvas.setFont('Helvetica-Bold', 14)
                    canvas.setFillColor(colors.HexColor('#2c3e50'))
                    canvas.drawCentredString(A4[0]/2, A4[1] - 50, tenant.name.upper())
            else:
                # No organization logo found, add organization name as header
                canvas.setFont('Helvetica-Bold', 14)
                canvas.setFillColor(colors.HexColor('#2c3e50'))
                canvas.drawCentredString(A4[0]/2, A4[1] - 50, tenant.name.upper())
            
            # Add watermark (Bossin or custom watermark text)
            watermark_text = tenant.theme.watermark_text if tenant.theme else 'Bossin'
            try:
                canvas.setFont('Helvetica', 60)
                canvas.setFillAlpha(0.1)  # 10% opacity
                canvas.setFillColor(colors.grey)
                # Rotate and position watermark diagonally
                canvas.rotate(45)
                canvas.drawCentredString(A4[0]/2, A4[1]/2, watermark_text)
                canvas.rotate(-45)
                canvas.setFillAlpha(1.0)  # Reset opacity
            except:
                pass  # Watermark is optional

            # Add page number
            canvas.setFont('Helvetica', 9)
            canvas.setFillColor(colors.black)
            canvas.drawRightString(A4[0] - 72, 30, f"Page {doc.page}")

            # Add generation date
            canvas.drawString(72, 30, f"Generated: {date.today().strftime('%B %d, %Y')}")

            canvas.restoreState()


        # Build PDF with custom page template
        doc.build(elements, onFirstPage=add_logo_and_header, onLaterPages=add_logo_and_header)

        # Create response
        pdf_data = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf_data, content_type='application/pdf')

        # Dynamic filename based on filters - Include organization name
        org_name_slug = tenant.slug.replace('-', '_')
        filename_parts = [org_name_slug, 'members_report']
        if filter_status:
            filename_parts.append(filter_status)
        if search_query:
            filename_parts.append(f'search_{search_query.replace(" ", "_")}')
        filename_parts.append(date.today().strftime('%Y%m%d'))

        filename = f"{'_'.join(filename_parts)}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        messages.error(request, f"Error exporting PDF: {str(e)}")
        return redirect_to_dashboard(request)
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def update_transaction_ajax(request):
    """Update transaction amount and/or note via AJAX - ORGANIZATION DATA ONLY"""
    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return JsonResponse({'success': False, 'error': 'Organization not found'})
        
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id')
        new_amount = data.get('amount')
        new_note = data.get('note', '')

        # Get the transaction - MUST belong to this organization
        transaction = get_object_or_404(Transaction, id=transaction_id, organization=tenant)
        
        # Store old amount for recalculation
        old_amount = transaction.amount

        # Validate and update amount if provided
        if new_amount is not None:
            try:
                new_amount = Decimal(str(new_amount))

                # Validate amount
                if new_amount <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'Amount must be greater than zero.'
                    })

                if new_amount > 10000000:  # 10 million TZS limit
                    return JsonResponse({
                        'success': False,
                        'error': 'Amount cannot exceed 10,000,000 TZS.'
                    })

                transaction.amount = new_amount

            except (InvalidOperation, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid amount format.'
                })

        # Update note
        if len(new_note.strip()) > 500:
            return JsonResponse({
                'success': False,
                'error': 'Note cannot exceed 500 characters.'
            })

        transaction.note = new_note.strip()
        transaction.save()

        # Recalculate member totals if amount changed
        if new_amount is not None and old_amount != new_amount:
            member = transaction.member
            # Recalculate paid_total from all transactions
            total_paid = member.transaction_set.aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')
            member.paid_total = total_paid
            member.save()

        return JsonResponse({
            'success': True,
            'new_amount': float(transaction.amount),
            'new_note': transaction.note,
            'formatted_amount': f"TZS {transaction.amount:,.0f}"
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating the transaction.'
        })

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def delete_transaction_ajax(request):
    """Delete transaction via AJAX - ORGANIZATION DATA ONLY"""
    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return JsonResponse({'success': False, 'error': 'Organization not found'})
        
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id')

        # Get the transaction - MUST belong to this organization
        transaction = get_object_or_404(Transaction, id=transaction_id, organization=tenant)
        member = transaction.member

        # Delete the transaction
        transaction.delete()

        # Recalculate member totals
        total_paid = member.transaction_set.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        member.paid_total = total_paid
        member.save()

        return JsonResponse({
            'success': True,
            'message': 'Transaction deleted successfully.'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while deleting the transaction.'
        })

# Your updated member_detail view
@login_required
def member_detail(request, member_id, org_slug=None):
    """Display member details and transactions"""
    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            messages.error(request, 'Organization not found')
            return redirect('tracker:dashboard')

        # Get member from THIS ORGANIZATION ONLY
        member = get_object_or_404(Member, id=member_id, organization=tenant)
        transactions = member.transaction_set.all().order_by('-date', '-id')

        # Check if user can add payments (admin or owner only)
        from .permissions import is_org_admin
        can_add_payment = is_org_admin(request.user, tenant)  # is_org_admin already includes owner

        if request.method == 'POST':
            # Check permission before processing POST
            if not can_add_payment:
                messages.error(request, 'You do not have permission to add payments.')
                return redirect('tracker:member_detail', member_id=member.id, org_slug=tenant.slug)
            
            transaction_form = TransactionForm(request.POST)
            if transaction_form.is_valid():
                try:
                    # Get the amount and validate it
                    amount = transaction_form.cleaned_data['amount']
                    # Validate amount is positive
                    if amount <= 0:
                        messages.error(request, 'Payment amount must be greater than zero.')
                        transaction_form = TransactionForm(request.POST)  # Re-populate form
                    else:
                        # Create the transaction
                        transaction = transaction_form.save(commit=False)
                        transaction.member = member
                        transaction.added_by = request.user
                        # Set date to today if not provided
                        if not transaction.date:
                            from datetime import date
                            transaction.date = date.today()
                        # Save the transaction
                        transaction.save()
                        # Update member's paid_total - use select_for_update to prevent race conditions
                        try:
                            from django.db import transaction as db_transaction
                            with db_transaction.atomic():
                                # Lock the member row for update
                                member = Member.objects.select_for_update().get(pk=member.pk)
                                # Recalculate paid_total from all transactions
                                total_paid = member.transaction_set.aggregate(
                                    total=Sum('amount')
                                )['total'] or Decimal('0.00')
                                member.paid_total = total_paid
                                member.save(update_fields=['paid_total'])
                            messages.success(request, f'Payment of TZS {amount:,.0f} recorded successfully!')
                            return redirect('tracker:member_detail', member_id=member.id, org_slug=tenant.slug)
                        except Exception as e:
                            # If updating paid_total fails, still save the transaction
                            messages.warning(request, f'Payment recorded but there was an issue updating member totals. Please contact administrator. Error: {str(e)}')
                            return redirect('tracker:member_detail', member_id=member.id, org_slug=tenant.slug)
                except Exception as e:
                    messages.error(request, f'Error recording payment: {str(e)}')
                    transaction_form = TransactionForm(request.POST)  # Re-populate form
            else:
                # Form validation errors
                for field, errors in transaction_form.errors.items():
                    for error in errors:
                        field_name = transaction_form.fields[field].label or field
                        messages.error(request, f'{field_name}: {error}')
        else:
            transaction_form = TransactionForm()

        context = {
            'member': member,
            'transactions': transactions,
            'transaction_form': transaction_form,
            'member_form': MemberForm(instance=member),
            'can_edit': request.user.is_staff,  # Simple check
            'can_add_payment': can_add_payment,  # Check for admin/owner
        }
        return render(request, 'tracker/member_detail.html', context)

    except Exception as e:
        messages.error(request, f'Error loading member details: {str(e)}')
        return redirect_to_dashboard(request)

def login_view(request):
    """Custom login view with organization membership check"""
    if request.user.is_authenticated:
        # Redirect to user's first organization dashboard
        from .models import OrganizationUser
        org_user = OrganizationUser.objects.filter(user=request.user, is_active=True).first()
        if org_user:
            return redirect('tracker:dashboard', org_slug=org_user.organization.slug)
        else:
            # User is authenticated but not in any active organization
            messages.error(request, 'Your account is not active in any organization. Please contact your administrator.')
            logout(request)
            return render(request, 'tracker/login.html', {'form': CustomLoginForm()})

    if request.method == 'POST':
        form = CustomLoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                # Check if user has active organization membership
                from .models import OrganizationUser
                org_user = OrganizationUser.objects.filter(user=user, is_active=True).first()
                
                if org_user:
                    login(request, user)
                    messages.success(request, f'Welcome back, {user.username}!')
                    return redirect('tracker:dashboard', org_slug=org_user.organization.slug)
                else:
                    # User exists but has no active organization
                    messages.error(request, 'Your account is not active in any organization. Please contact your administrator.')
            else:
                messages.error(request, 'Invalid username or password.')
    else:
        form = CustomLoginForm()

    return render(request, 'tracker/login.html', {'form': form})


@login_required
def dashboard(request, org_slug=None):
    """Main dashboard with member table and statistics"""
    # Get filter parameters
    search_query = request.GET.get('search', '')
    filter_status = request.GET.get('filter', '')

    # Initialize default values
    total_collected = Decimal('0.00')
    total_pledged = Decimal('0.00')
    target_amount = Decimal('210000.00')  # Default target
    progress_percentage = 0
    not_paid_count = 0
    incomplete_count = 0
    complete_count = 0
    exceeded_count = 0
    page_obj = None
    members = []

    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            messages.error(request, 'Organization not found')
            return redirect('tracker:dashboard')
        
        # Get organization's target amount from theme
        try:
            target_amount = Decimal(str(tenant.theme.target_amount))
        except:
            target_amount = Decimal('210000.00')

        # Get members for THIS ORGANIZATION ONLY with individual error handling
        members = []
        try:
            # Try to get members for this organization
            all_members = Member.objects.filter(organization=tenant)
            for member in all_members:
                try:
                    # Verify each member's decimal fields are valid
                    if member.pledge is None or member.paid_total is None:
                        # Fix invalid values
                        if member.pledge is None:
                            member.pledge = Decimal('70000.00')
                        if member.paid_total is None:
                            member.paid_total = Decimal('0.00')
                        member.save()

                    members.append(member)
                except Exception as e:
                    # Skip problematic members
                    continue
        except Exception as e:
            # If bulk query fails, get members one by one
            try:
                member_ids = Member.objects.filter(organization=tenant).values_list('id', flat=True)
                for member_id in member_ids:
                    try:
                        member = Member.objects.get(id=member_id, organization=tenant)
                        # Fix any invalid decimal values
                        if member.pledge is None:
                            member.pledge = Decimal('70000.00')
                        if member.paid_total is None:
                            member.paid_total = Decimal('0.00')
                        member.save()
                        members.append(member)
                    except:
                        continue
            except:
                members = []

        # Apply search filter
        if search_query and members:
            filtered_members = []
            for member in members:
                try:
                    if (search_query.lower() in member.name.lower() or
                        (member.phone and search_query in member.phone) or
                        (member.email and search_query.lower() in member.email.lower())):
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        # Apply status filter
        if filter_status and members:
            filtered_members = []
            for member in members:
                try:
                    if filter_status == 'incomplete' and member.paid_total > 0 and member.paid_total < member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'complete' and member.paid_total >= member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'pledged' and member.pledge > 70000:
                        filtered_members.append(member)
                    elif filter_status == 'not_started' and member.paid_total == 0:
                        filtered_members.append(member)
                    elif filter_status == 'exceeded' and member.paid_total > member.pledge:
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        # Calculate statistics safely
        try:
            # Manual calculation to avoid aggregation issues
            total_collected = Decimal('0.00')
            total_pledged = Decimal('0.00')
            not_paid_count = 0
            incomplete_count = 0
            complete_count = 0
            exceeded_count = 0

            for member in members:
                try:
                    # Ensure we have valid decimal values
                    pledge = member.pledge if member.pledge is not None else Decimal('70000.00')
                    paid_total = member.paid_total if member.paid_total is not None else Decimal('0.00')

                    total_pledged += pledge
                    total_collected += paid_total

                    # Count by status
                    if paid_total == 0:
                        not_paid_count += 1
                    elif paid_total < pledge:
                        incomplete_count += 1
                    elif paid_total == pledge:
                        complete_count += 1
                    else:  # paid_total > pledge
                        exceeded_count += 1
                except:
                    # Skip problematic members
                    continue

        except Exception as e:
            # If calculation fails, use default values
            total_collected = Decimal('0.00')
            total_pledged = Decimal('0.00')
            not_paid_count = 0
            incomplete_count = 0
            complete_count = 0
            exceeded_count = 0

        # Calculate progress percentage
        progress_percentage = (total_collected / target_amount * 100) if target_amount > 0 else 0

    except Exception as e:
        # If everything fails, use default values
        pass

    context = {
        'members': members,
        'search_query': search_query,
        'filter_status': filter_status,
        'total_collected': total_collected,
        'total_pledged': total_pledged,
        'target_amount': target_amount,
        'progress_percentage': progress_percentage,
        'not_paid_count': not_paid_count,
        'incomplete_count': incomplete_count,
        'complete_count': complete_count,
        'exceeded_count': exceeded_count,
        'quick_member_form': QuickMemberForm(),
        'excel_import_form': ExcelImportForm(),
        'can_edit': request.user.is_staff,  # Simple check
    }

    return render(request, 'tracker/dashboard.html', context)

# Updated views.py - Add these new AJAX views and update your member_detail view


@login_required
def add_member(request, org_slug=None):
    """Add new member - ORGANIZATION DATA ONLY"""
    # Get tenant from request for data isolation
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect_to_dashboard(request)
    
    if request.method == 'POST':
        form = MemberForm(request.POST)
        if form.is_valid():
            member = form.save(commit=False)
            member.organization = tenant  # Assign to current organization
            member.save()
            messages.success(request, f'Member "{member.name}" added successfully!')
            return redirect('tracker:member_detail', member_id=member.id, org_slug=org_slug)
    else:
        form = MemberForm()

    return render(request, 'tracker/add_member.html', {'form': form})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def update_transaction_ajax(request):
    """Update transaction amount and/or note via AJAX"""
    try:
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id')
        new_amount = data.get('amount')
        new_note = data.get('note', '')

        # Get the transaction
        transaction = get_object_or_404(Transaction, id=transaction_id)

        # Store old amount for recalculation
        old_amount = transaction.amount

        # Validate and update amount if provided
        if new_amount is not None:
            try:
                new_amount = Decimal(str(new_amount))

                # Validate amount
                if new_amount <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'Amount must be greater than zero.'
                    })

                if new_amount > 10000000:  # 10 million TZS limit
                    return JsonResponse({
                        'success': False,
                        'error': 'Amount cannot exceed 10,000,000 TZS.'
                    })

                transaction.amount = new_amount

            except (InvalidOperation, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid amount format.'
                })

        # Update note
        if len(new_note.strip()) > 500:
            return JsonResponse({
                'success': False,
                'error': 'Note cannot exceed 500 characters.'
            })

        transaction.note = new_note.strip()
        transaction.save()

        # Recalculate member totals if amount changed
        if new_amount is not None and old_amount != new_amount:
            member = transaction.member
            # You might want to add a method to recalculate totals
            # member.recalculate_totals()

        return JsonResponse({
            'success': True,
            'new_amount': float(transaction.amount),
            'new_note': transaction.note,
            'formatted_amount': f"TZS {transaction.amount:,.0f}"
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating the transaction.'
        })

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def delete_transaction_ajax(request):
    """Delete transaction via AJAX"""
    try:
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id')

        # Get the transaction
        transaction = get_object_or_404(Transaction, id=transaction_id)
        member = transaction.member

        # Delete the transaction
        transaction.delete()

        # Recalculate member totals
        # member.recalculate_totals()

        return JsonResponse({
            'success': True,
            'message': 'Transaction deleted successfully.'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while deleting the transaction.'
        })

# Update your existing edit_member view to include the transaction form
@login_required
def edit_member(request, member_id, org_slug=None):
    """Edit member details - ORGANIZATION DATA ONLY"""
    # Get tenant from request for data isolation
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect_to_dashboard(request)
    
    # Get member from THIS ORGANIZATION ONLY
    member = get_object_or_404(Member, id=member_id, organization=tenant)

    if request.method == 'POST':
        form = MemberForm(request.POST, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, f'Member "{member.name}" updated successfully!')
            return redirect('tracker:member_detail', member_id=member.id, org_slug=org_slug)
    else:
        form = MemberForm(instance=member)

    return render(request, 'tracker/edit_member.html', {'form': form, 'member': member})

from django.contrib import messages
from decimal import Decimal, InvalidOperation
from io import BytesIO
import openpyxl
from datetime import date
from .models import Member, Transaction
from .forms import ExcelImportForm

@login_required
def import_excel(request, org_slug=None):
    """Import members and payment transactions from an Excel file - ORGANIZATION DATA ONLY"""
    # Get tenant from request for data isolation
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect_to_dashboard(request)
    
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            update_existing = form.cleaned_data['update_existing']
            default_pledge = form.cleaned_data['default_pledge']

            try:
                workbook = openpyxl.load_workbook(BytesIO(excel_file.read()))
                sheet = workbook.active

                created_count = 0
                updated_count = 0
                transaction_count = 0
                errors = []

                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not row[0]:
                            continue  # Skip rows without a name

                        name = str(row[0]).strip()
                        pledge = row[1]
                        paid = row[2] if len(row) > 2 else None
                        phone = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                        email = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                        course = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        year = str(row[6]).strip() if len(row) > 6 and row[6] else ''

                        # Pledge
                        try:
                            pledge = Decimal(str(pledge)) if pledge else default_pledge
                        except InvalidOperation:
                            pledge = default_pledge
                            errors.append(f"Row {idx}: Invalid pledge, using default for '{name}'")

                        # Paid
                        try:
                            paid = Decimal(str(paid)) if paid else Decimal('0.00')
                        except InvalidOperation:
                            paid = Decimal('0.00')
                            errors.append(f"Row {idx}: Invalid paid value for '{name}', using 0.00")

                        member_created = False
                        if update_existing:
                            member, member_created = Member.objects.update_or_create(
                                organization=tenant,
                                name=name,
                                defaults={
                                    'pledge': pledge,
                                    'phone': phone or None,
                                    'email': email or None,
                                    'course': course or None,
                                    'year': year or None,
                                }
                            )
                        else:
                            member = Member.objects.filter(organization=tenant, name=name).first()
                            if not member:
                                member = Member.objects.create(
                                    organization=tenant,
                                    name=name,
                                    pledge=pledge,
                                    phone=phone or None,
                                    email=email or None,
                                    course=course or None,
                                    year=year or None,
                                )
                                member_created = True
                            else:
                                errors.append(f"Row {idx}: Member '{name}' already exists and update is off.")
                                continue

                        if member_created:
                            created_count += 1
                        else:
                            updated_count += 1

                        # Transaction
                        if paid > 0:
                            existing_txn = Transaction.objects.filter(
                                organization=tenant,
                                member=member,
                                date=date.today(),
                                added_by=request.user,
                                note__icontains="Imported via Excel"
                            ).first()

                            if existing_txn:
                                # Update the amount if different
                                if existing_txn.amount != paid:
                                    existing_txn.amount = paid
                                    existing_txn.note = f"Updated via Excel on {date.today().isoformat()}"
                                    existing_txn.save()
                                    transaction_count += 1
                            else:
                                # Create new transaction
                                Transaction.objects.create(
                                    organization=tenant,
                                    member=member,
                                    amount=paid,
                                    date=date.today(),
                                    added_by=request.user,
                                    note=f"Imported via Excel on {date.today().isoformat()}"
                                )
                                transaction_count += 1


                    except Exception as e:
                        errors.append(f"Row {idx}: Error processing member '{row[0]}' - {str(e)}")

                if created_count:
                    messages.success(request, f"✅ Created {created_count} new members.")
                if updated_count:
                    messages.info(request, f"🔄 Updated {updated_count} existing members.")
                if transaction_count:
                    messages.success(request, f"💰 Recorded {transaction_count} payments.")

                if errors:
                    for err in errors[:5]:
                        messages.warning(request, f"⚠️ {err}")
                    if len(errors) > 5:
                        messages.warning(request, f"...and {len(errors) - 5} more issues.")

                return redirect_to_dashboard(request)

            except Exception as e:
                messages.error(request, f"📄 Excel reading error: {str(e)}")

        else:
            messages.error(request, "⚠️ Invalid form submission.")
    else:
        form = ExcelImportForm()

    return render(request, 'tracker/import_excel.html', {'form': form})



@login_required
def admin_log(request, org_slug=None):
    """Admin log showing all transactions"""
    # Get tenant from request for data isolation
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect('tracker:dashboard')

    # Get transactions for THIS ORGANIZATION ONLY
    transactions = Transaction.objects.select_related('member', 'added_by').filter(member__organization=tenant)
    # Apply filters
    boss_filter = request.GET.get('boss', '')
    date_filter = request.GET.get('date', '')
    amount_filter = request.GET.get('amount', '')

    if boss_filter:
        transactions = transactions.filter(added_by__username__icontains=boss_filter)
    if date_filter:
        transactions = transactions.filter(date=date_filter)
    if amount_filter:
        try:
            amount = Decimal(amount_filter)
            transactions = transactions.filter(amount=amount)
        except:
            pass

    # Pagination
    paginator = Paginator(transactions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get member edit logs (owner only)
    member_edit_logs = []
    is_owner = is_org_owner(request.user, tenant)
    if is_owner:
        member_edit_logs = MemberEditLog.objects.select_related('member', 'edited_by').filter(
            organization=tenant
        ).order_by('-created_at')
        
        # Apply filters for member edit logs
        edit_member_filter = request.GET.get('edit_member', '')
        edit_field_filter = request.GET.get('edit_field', '')
        
        if edit_member_filter:
            member_edit_logs = member_edit_logs.filter(member__name__icontains=edit_member_filter)
        if edit_field_filter:
            member_edit_logs = member_edit_logs.filter(field_changed=edit_field_filter)
        
        # Pagination for member edit logs
        edit_paginator = Paginator(member_edit_logs, 25)
        edit_page_number = request.GET.get('edit_page')
        member_edit_logs = edit_paginator.get_page(edit_page_number)

    context = {
        'page_obj': page_obj,
        'transactions': page_obj,
        'boss_filter': boss_filter,
        'date_filter': date_filter,
        'amount_filter': amount_filter,
        'can_edit': request.user.is_staff,  # Simple check
        'member_edit_logs': member_edit_logs,
        'is_owner': is_owner,
        'edit_member_filter': edit_member_filter if is_owner else '',
        'edit_field_filter': edit_field_filter if is_owner else '',
    }

    return render(request, 'tracker/admin_log.html', context)


# AJAX Views
@login_required
@require_POST
@csrf_exempt
def update_member_ajax(request):
    """AJAX endpoint for updating member data"""
    try:
        data = json.loads(request.body)
        member_id = data.get('member_id')
        field = data.get('field')
        value = data.get('value')

        member = get_object_or_404(Member, id=member_id)

        if field in ['pledge', 'paid_total']:
            try:
                # Clean the value and convert to Decimal
                if value is None or value == '':
                    value = Decimal('0.00')
                else:
                    # Remove any non-numeric characters except decimal point
                    cleaned_value = str(value).replace(',', '').strip()
                    value = Decimal(cleaned_value)

                # Ensure value is not negative
                if value < 0:
                    value = Decimal('0.00')

                setattr(member, field, value)
                member.save()

                # Return updated member data
                return JsonResponse({
                    'success': True,
                    'pledge': float(member.pledge),
                    'paid_total': float(member.paid_total),
                    'remaining': float(member.remaining),
                    'is_complete': member.is_complete,
                    'is_incomplete': member.is_incomplete,
                    'not_started': member.not_started,
                    'has_exceeded': member.has_exceeded,
                    'status_display': member.status_display,
                })
            except (ValueError, TypeError, InvalidOperation) as e:
                return JsonResponse({'success': False, 'error': f'Invalid value: {str(e)}'})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid field'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
@csrf_exempt
def add_member_ajax(request):
    """AJAX endpoint for quick member addition"""
    try:
        data = json.loads(request.body)
        name = data.get('name')
        pledge = data.get('pledge', 70000)

        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})

        # Check if member already exists
        if Member.objects.filter(name__iexact=name).exists():
            return JsonResponse({'success': False, 'error': 'Member already exists'})

        member = Member.objects.create(
            name=name,
            pledge=Decimal(pledge)
        )

        # Return the new member data
        return JsonResponse({
            'success': True,
            'member': {
                'id': member.id,
                'name': member.name,
                'pledge': float(member.pledge),
                'paid_total': float(member.paid_total),
                'remaining': float(member.remaining),
                'status_display': member.status_display,
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def update_transaction_note_ajax(request, org_slug=None):
    """AJAX endpoint for updating transaction notes - ORGANIZATION DATA ONLY"""
    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return JsonResponse({'success': False, 'error': 'Organization not found'})
        
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id')
        note = data.get('note', '')

        # Get the transaction - MUST belong to this organization
        transaction = get_object_or_404(Transaction, id=transaction_id, organization=tenant)
        transaction.note = note
        transaction.save()

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})



@org_admin_required
def daily_collection(request, org_slug=None):
    """Daily collection page for recording payments"""
    # Get filter parameters
    search_query = request.GET.get('search', '')
    filter_status = request.GET.get('filter', '')

    # Initialize default values
    total_collected = Decimal('0.00')
    total_pledged = Decimal('0.00')
    target_amount = Decimal('210000.00')  # Default target
    progress_percentage = 0
    not_paid_count = 0
    incomplete_count = 0
    complete_count = 0
    exceeded_count = 0
    page_obj = None
    members = []

    try:
        # Get tenant from request for data isolation
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            messages.error(request, 'Organization not found')
            return redirect('tracker:dashboard')
        
        # Get organization's target amount from theme
        try:
            target_amount = Decimal(str(tenant.theme.target_amount))
        except:
            target_amount = Decimal('210000.00')

        # Get members for THIS ORGANIZATION ONLY with individual error handling
        members = []
        try:
            # Try to get members for this organization
            all_members = Member.objects.filter(organization=tenant)
            for member in all_members:
                try:
                    # Verify each member's decimal fields are valid
                    if member.pledge is None or member.paid_total is None:
                        # Fix invalid values
                        if member.pledge is None:
                            member.pledge = Decimal('70000.00')
                        if member.paid_total is None:
                            member.paid_total = Decimal('0.00')
                        member.save()

                    members.append(member)
                except Exception as e:
                    # Skip problematic members
                    continue
        except Exception as e:
            # If bulk query fails, get members one by one
            try:
                member_ids = Member.objects.filter(organization=tenant).values_list('id', flat=True)
                for member_id in member_ids:
                    try:
                        member = Member.objects.get(id=member_id, organization=tenant)
                        # Fix any invalid decimal values
                        if member.pledge is None:
                            member.pledge = Decimal('70000.00')
                        if member.paid_total is None:
                            member.paid_total = Decimal('0.00')
                        member.save()
                        members.append(member)
                    except:
                        continue
            except:
                members = []

        # Apply search filter
        if search_query and members:
            filtered_members = []
            for member in members:
                try:
                    if (search_query.lower() in member.name.lower() or
                        (member.phone and search_query in member.phone) or
                        (member.email and search_query.lower() in member.email.lower())):
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        # Apply status filter
        if filter_status and members:
            filtered_members = []
            for member in members:
                try:
                    if filter_status == 'incomplete' and member.paid_total > 0 and member.paid_total < member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'complete' and member.paid_total >= member.pledge:
                        filtered_members.append(member)
                    elif filter_status == 'pledged' and member.pledge > 70000:
                        filtered_members.append(member)
                    elif filter_status == 'not_started' and member.paid_total == 0:
                        filtered_members.append(member)
                    elif filter_status == 'exceeded' and member.paid_total > member.pledge:
                        filtered_members.append(member)
                except:
                    continue
            members = filtered_members

        # Calculate statistics safely
        try:
            # Manual calculation to avoid aggregation issues
            total_collected = Decimal('0.00')
            total_pledged = Decimal('0.00')
            not_paid_count = 0
            incomplete_count = 0
            complete_count = 0
            exceeded_count = 0

            for member in members:
                try:
                    # Ensure we have valid decimal values
                    pledge = member.pledge if member.pledge is not None else Decimal('70000.00')
                    paid_total = member.paid_total if member.paid_total is not None else Decimal('0.00')

                    total_pledged += pledge
                    total_collected += paid_total

                    # Count by status
                    if paid_total == 0:
                        not_paid_count += 1
                    elif paid_total < pledge:
                        incomplete_count += 1
                    elif paid_total == pledge:
                        complete_count += 1
                    else:  # paid_total > pledge
                        exceeded_count += 1
                except:
                    # Skip problematic members
                    continue

        except Exception as e:
            # If calculation fails, use default values
            total_collected = Decimal('0.00')
            total_pledged = Decimal('0.00')
            not_paid_count = 0
            incomplete_count = 0
            complete_count = 0
            exceeded_count = 0

        # Calculate progress percentage
        progress_percentage = (total_collected / target_amount * 100) if target_amount > 0 else 0

        # Pagination with error handling
        try:
            # Create a simple list-based pagination to avoid queryset issues
            page_size = 20
            page_number = request.GET.get('page', 1)
            try:
                page_number = int(page_number)
            except:
                page_number = 1

            # Calculate pagination manually
            total_members = len(members)
            total_pages = (total_members + page_size - 1) // page_size

            # Ensure page number is valid
            if page_number < 1:
                page_number = 1
            elif page_number > total_pages and total_pages > 0:
                page_number = total_pages

            # Create a simple page object
            class SimplePage:
                def __init__(self, object_list, number, paginator):
                    self.paginator = paginator
                    self.number = number

                    # Calculate the slice for this page
                    start_index = (number - 1) * paginator.per_page
                    end_index = start_index + paginator.per_page
                    self.object_list = object_list[start_index:end_index]

                def has_other_pages(self):
                    return self.paginator.num_pages > 1

                def has_previous(self):
                    return self.number > 1

                def has_next(self):
                    return self.number < self.paginator.num_pages

                def previous_page_number(self):
                    return self.number - 1

                def next_page_number(self):
                    return self.number + 1

            class SimplePaginator:
                def __init__(self, object_list, per_page):
                    self.object_list = object_list
                    self.per_page = per_page
                    self.count = len(object_list)
                    self.num_pages = (self.count + per_page - 1) // per_page

                def get_page(self, number):
                    return SimplePage(self.object_list, number, self)

            paginator = SimplePaginator(members, page_size)
            page_obj = paginator.get_page(page_number)

        except Exception as e:
            # If pagination fails, create empty page
            page_obj = None
            members = []

    except Exception as e:
        # If everything fails, use default values
        pass

    context = {
        'page_obj': page_obj,
        'members': page_obj.object_list if page_obj else [],
        'search_query': search_query,
        'filter_status': filter_status,
        'total_collected': total_collected,
        'total_pledged': total_pledged,
        'target_amount': target_amount,
        'progress_percentage': progress_percentage,
        'not_paid_count': not_paid_count,
        'incomplete_count': incomplete_count,
        'complete_count': complete_count,
        'exceeded_count': exceeded_count,
        'today': timezone.now().date(),
        'can_edit': request.user.is_staff,  # Simple check
    }

    return render(request, 'tracker/daily_collection.html', context)


@org_staff_required
@require_POST
@csrf_exempt
def record_daily_payment_ajax(request, org_slug=None):
    """AJAX endpoint for recording daily payments"""
    try:
        data = json.loads(request.body)
        member_id = data.get('member_id')
        payment_amount = data.get('payment_amount')
        note = data.get('note', '')

        if not member_id or not payment_amount:
            return JsonResponse({'success': False, 'error': 'Member ID and payment amount are required'})

        # Get member
        member = get_object_or_404(Member, id=member_id)

        # Convert payment amount to decimal
        try:
            payment_amount = Decimal(str(payment_amount).replace(',', '').strip())
            if payment_amount <= 0:
                return JsonResponse({'success': False, 'error': 'Payment amount must be greater than 0'})
        except (ValueError, InvalidOperation):
            return JsonResponse({'success': False, 'error': 'Invalid payment amount'})

        # Create transaction
        transaction = Transaction.objects.create(
            member=member,
            amount=payment_amount,
            date=timezone.now().date(),
            added_by=request.user,
            note=note
        )

        # Update member's paid_total (this happens automatically via the model's save method)
        member.refresh_from_db()

        # Return updated member data
        return JsonResponse({
            'success': True,
            'transaction_id': transaction.id,
            'member': {
                'id': member.id,
                'name': member.name,
                'pledge': float(member.pledge),
                'paid_total': float(member.paid_total),
                'remaining': float(member.remaining),
                'is_complete': member.is_complete,
                'is_incomplete': member.is_incomplete,
                'not_started': member.not_started,
                'has_exceeded': member.has_exceeded,
                'status_display': member.status_display,
            },
            'message': f'Payment of TZS {payment_amount:,.0f} recorded successfully!'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ============================================================================
# ORGANIZATION ADMIN DASHBOARD VIEWS
# ============================================================================

from .permissions import org_admin_required
from .forms import OrganizationSettingsForm, OrganizationThemeForm, OrganizationUserForm, AddOrganizationUserForm
from .models import PaymentRequest

@org_admin_required
def org_admin_dashboard(request, org_slug=None):
    """Organization admin dashboard - main page"""
    tenant = request.tenant
    
    # Get organization stats
    members_count = Member.objects.filter(organization=tenant).count()
    staff_count = OrganizationUser.objects.filter(organization=tenant, is_active=True).count()
    transactions_count = Transaction.objects.filter(organization=tenant).count()
    
    # Get total collected
    total_collected = Transaction.objects.filter(organization=tenant).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    context = {
        'organization': tenant,
        'members_count': members_count,
        'staff_count': staff_count,
        'transactions_count': transactions_count,
        'total_collected': total_collected,
    }
    
    return render(request, 'tracker/org_admin/dashboard.html', context)


@org_admin_required
def org_settings(request, org_slug=None):
    """Edit organization settings"""
    tenant = request.tenant
    
    if request.method == 'POST':
        form = OrganizationSettingsForm(request.POST, instance=tenant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Organization settings updated successfully!')
            return redirect('tracker:org_settings', org_slug=org_slug)
    else:
        form = OrganizationSettingsForm(instance=tenant)
    
    context = {
        'form': form,
        'organization': tenant,
    }
    
    return render(request, 'tracker/org_admin/settings.html', context)


@org_admin_required
def org_branding(request, org_slug=None):
    """Edit organization branding and theme"""
    tenant = request.tenant
    
    theme = tenant.theme
    
    if request.method == 'POST':
        form = OrganizationThemeForm(request.POST, request.FILES, instance=theme)
        if form.is_valid():
            form.save()
            messages.success(request, 'Branding and theme updated successfully!')
            return redirect('tracker:org_branding', org_slug=org_slug)
    else:
        form = OrganizationThemeForm(instance=theme)
    
    context = {
        'form': form,
        'organization': tenant,
        'theme': theme,
    }
    
    return render(request, 'tracker/org_admin/branding.html', context)


@org_admin_required
def org_billing(request, org_slug=None):
    """Owner/Admin billing page inside admin area."""
    tenant = request.tenant
    category = tenant.category
    category_discount = 50 if category == 'religious' else 35
    requests_qs = PaymentRequest.objects.filter(organization=tenant).order_by('-created_at')[:20]
    # Build status - check subscription_status first, then dates
    now = timezone.now()
    last_approved = PaymentRequest.objects.filter(organization=tenant, status='approved', is_trial=False).order_by('-created_at').first()
    
    # Check subscription status first - use subscription_status field as primary source
    if tenant.subscription_status == 'SUBSCRIBED':
        # Check if subscription is still active (not expired)
        if tenant.subscription_expires_at:
            active_sub = tenant.subscription_expires_at > now
            if active_sub:
                status_label = 'Subscribed'
                started_str = last_approved.created_at.strftime('%Y-%m-%d %H:%M') if last_approved else 'Unknown'
                months_str = str(last_approved.months) if last_approved else '—'
                amount_str = f"{last_approved.amount_tzs} TZS" if last_approved else '—'
                days_left = (tenant.subscription_expires_at - now).days
                status_lines = [
                    f"Started: {started_str}",
                    f"Duration: {months_str} month(s) | Amount: {amount_str}",
                    f"Ends: {tenant.subscription_expires_at.strftime('%Y-%m-%d %H:%M')} ({days_left} days left)"
                ]
            else:
                # Status is SUBSCRIBED but expired
                status_label = 'Not Subscribed'
                status_lines = [
                    f"Subscription expired: {tenant.subscription_expires_at.strftime('%Y-%m-%d %H:%M')}",
                    'Please submit a new subscription request.'
                ]
        else:
            # Status is SUBSCRIBED but expiration date is missing - this shouldn't happen
            # But we'll show Subscribed status anyway since that's what the database says
            status_label = 'Subscribed'
            status_lines = [
                'Subscription active (expiration date pending update)',
                'Please contact admin if this persists.'
            ]
    elif tenant.subscription_status == 'FREE_TRIAL':
        # Check if trial is still active
        trial_active = bool(tenant.trial_started_at and (tenant.trial_started_at + timezone.timedelta(days=7) > now))
        if trial_active:
            status_label = 'Trial'
            trial_end = tenant.trial_started_at + timezone.timedelta(days=7)
            days_left = (trial_end - now).days
            status_lines = [
                f"Trial ends: {trial_end.strftime('%Y-%m-%d %H:%M')} ({days_left} days left)"
            ]
        else:
            status_label = 'Not Subscribed'
            status_lines = [
                'Trial has expired. Please submit a subscription request.'
            ]
    else:
        # NOT_SUBSCRIBED
        status_label = 'Not Subscribed'
        status_lines = [
            'No active subscription. Please submit a subscription request.'
        ]
    context = {
        'organization': tenant,
        'requests': requests_qs,
        'base_price': 19800,
        'category_discount': category_discount,
        'status_label': status_label,
        'status_lines': status_lines,
    }
    return render(request, 'tracker/admin/billing.html', context)


@org_admin_required
def org_financial_settings(request, org_slug=None):
    """Edit organization financial settings (default pledge and target amount)"""
    tenant = request.tenant
    theme = tenant.theme
    
    if request.method == 'POST':
        try:
            default_pledge = Decimal(request.POST.get('default_pledge_amount', 0))
            target_amount = Decimal(request.POST.get('target_amount', 0))
            
            if default_pledge < 0 or target_amount < 0:
                messages.error(request, 'Amounts must be positive numbers.')
            else:
                theme.default_pledge_amount = default_pledge
                theme.target_amount = target_amount
                theme.save()
                messages.success(request, 'Financial settings updated successfully!')
                return redirect('tracker:org_financial_settings', org_slug=org_slug)
        except (ValueError, InvalidOperation):
            messages.error(request, 'Please enter valid numbers for the amounts.')
    
    context = {
        'organization': tenant,
        'theme': theme,
        'default_pledge_amount': float(theme.default_pledge_amount),
        'target_amount': float(theme.target_amount),
    }
    
    return render(request, 'tracker/org_admin/financial_settings.html', context)


@org_admin_required
def org_staff_management(request, org_slug=None):
    """Manage organization staff and users"""
    tenant = request.tenant
    
    # Get all staff members
    staff_members = OrganizationUser.objects.filter(organization=tenant).select_related('user')
    
    context = {
        'organization': tenant,
        'staff_members': staff_members,
    }
    
    return render(request, 'tracker/org_admin/staff_management.html', context)


@org_admin_required
def add_staff_member(request, org_slug=None):
    """Add new staff member to organization - creates user on-the-fly"""
    tenant = request.tenant
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        role = request.POST.get('role', 'staff')
        
        if not username or not password:
            messages.error(request, 'Username and password are required.')
            return redirect('tracker:add_staff_member', org_slug=org_slug)
        
        try:
            # Check if user already exists in organization
            if OrganizationUser.objects.filter(user__username=username, organization=tenant).exists():
                messages.error(request, f'User "{username}" is already a member of this organization.')
                return redirect('tracker:add_staff_member', org_slug=org_slug)
            
            # Create user
            user = User.objects.create_user(
                username=username,
                password=password
            )
            
            # Create organization user
            OrganizationUser.objects.create(
                user=user,
                organization=tenant,
                role=role,
                is_active=True
            )
            
            messages.success(request, f'Staff member "{username}" added successfully with role "{role}"!')
            return redirect('tracker:org_staff_management', org_slug=org_slug)
                    
        except Exception as e:
            messages.error(request, f'Error adding staff member: {str(e)}')
            return redirect('tracker:add_staff_member', org_slug=org_slug)
    
    staff_members = tenant.staff_members.all()
    context = {
        'organization': tenant,
        'staff_members': staff_members,
    }
    
    return render(request, 'tracker/org_admin/add_staff_new.html', context)


@org_admin_required
def edit_staff_member(request, org_slug=None, staff_id=None):
    """Edit staff member role and status"""
    tenant = request.tenant
    
    org_user = get_object_or_404(OrganizationUser, id=staff_id, organization=tenant)
    
    if request.method == 'POST':
        form = OrganizationUserForm(request.POST, instance=org_user)
        if form.is_valid():
            form.save()
            messages.success(request, f'User "{org_user.user.username}" updated successfully!')
            return redirect('tracker:org_staff_management', org_slug=org_slug)
    else:
        form = OrganizationUserForm(instance=org_user)
    
    # Hide owner role option for non-owners
    is_owner = is_org_owner(request.user, tenant)
    if not is_owner:
        # Remove 'owner' option from role choices
        form.fields['role'].choices = [
            (choice[0], choice[1]) for choice in form.fields['role'].choices 
            if choice[0] != 'owner'
        ]
    
    context = {
        'form': form,
        'organization': tenant,
        'org_user': org_user,
    }
    
    return render(request, 'tracker/org_admin/edit_staff.html', context)


@org_admin_required
@require_POST
def remove_staff_member(request, org_slug=None, staff_id=None):
    """Remove staff member from organization"""
    tenant = request.tenant
    
    org_user = get_object_or_404(OrganizationUser, id=staff_id, organization=tenant)
    username = org_user.user.username
    
    # Prevent removing the last owner
    owner_count = OrganizationUser.objects.filter(
        organization=tenant,
        role='owner',
        is_active=True
    ).count()
    
    if org_user.role == 'owner' and owner_count <= 1:
        messages.error(request, 'Cannot remove the last owner of the organization.')
    else:
        org_user.delete()
        messages.success(request, f'User "{username}" removed from organization.')
    
    return redirect('tracker:org_staff_management', org_slug=org_slug)


def offline_view(request):
    """Offline page for PWA"""
    return render(request, 'offline.html')


# ============================================================================
# ONBOARDING FLOW VIEWS
# ============================================================================

@login_required
def onboarding_financial(request, org_slug):
    """Step 1: Financial Settings"""
    try:
        organization = Organization.objects.get(slug=org_slug)
        theme = organization.theme
    except Organization.DoesNotExist:
        messages.error(request, 'Organization not found.')
        return redirect('tracker:landing')
    
    if request.method == 'POST':
        default_pledge = request.POST.get('default_pledge_amount', theme.default_pledge_amount)
        target_amount = request.POST.get('target_amount', theme.target_amount)
        
        try:
            theme.default_pledge_amount = Decimal(default_pledge)
            theme.target_amount = Decimal(target_amount)
            theme.save()
            messages.success(request, 'Financial settings saved!')
        except Exception as e:
            messages.error(request, f'Error saving settings: {str(e)}')
        
        # Redirect to next step
        return redirect('tracker:onboarding_branding', org_slug=org_slug)
    
    context = {
        'organization': organization,
        'theme': theme,
        'step': 1,
        'total_steps': 5
    }
    return render(request, 'tracker/onboarding/financial.html', context)


@login_required
def onboarding_branding(request, org_slug):
    """Step 2: Organization Branding & Theme"""
    try:
        organization = Organization.objects.get(slug=org_slug)
        theme = organization.theme
    except Organization.DoesNotExist:
        messages.error(request, 'Organization not found.')
        return redirect('tracker:landing')
    
    if request.method == 'POST':
        try:
            # Category selection (auto-sets theme defaults)
            selected_category = request.POST.get('category')
            if selected_category and selected_category in dict(Organization.CATEGORY_CHOICES):
                organization.category = selected_category
                organization.save(update_fields=['category'])
                # Apply default colors based on category
                from .models import apply_category_default_theme
                apply_category_default_theme(theme, selected_category)
            
            theme.navbar_title = request.POST.get('navbar_title', theme.navbar_title)
            theme.watermark_text = request.POST.get('watermark_text', theme.watermark_text)
            
            if 'logo' in request.FILES:
                theme.logo = request.FILES['logo']
            
            theme.save()
            messages.success(request, 'Branding settings saved!')
        except Exception as e:
            messages.error(request, f'Error saving branding: {str(e)}')
        
        # Redirect to next step
        return redirect('tracker:onboarding_staff', org_slug=org_slug)
    
    context = {
        'organization': organization,
        'theme': theme,
        'category_choices': Organization.CATEGORY_CHOICES,
        'step': 2,
        'total_steps': 5
    }
    return render(request, 'tracker/onboarding/branding.html', context)


@login_required
def onboarding_staff(request, org_slug):
    """Step 3: Staff Management"""
    try:
        organization = Organization.objects.get(slug=org_slug)
    except Organization.DoesNotExist:
        messages.error(request, 'Organization not found.')
        return redirect('tracker:landing')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_staff':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            role = request.POST.get('role', 'staff')
            
            if not username or not password:
                messages.error(request, 'Username and password are required.')
                return redirect('tracker:onboarding_staff', org_slug=org_slug)
            
            try:
                # Create user
                user = User.objects.create_user(
                    username=username,
                    password=password
                )
                
                # Add to organization
                OrganizationUser.objects.create(
                    user=user,
                    organization=organization,
                    role=role,
                    is_active=True
                )
                
                messages.success(request, f'Staff member "{username}" added successfully!')
            except Exception as e:
                messages.error(request, f'Error adding staff: {str(e)}')
        
        # Redirect to next step
        return redirect('tracker:onboarding_import', org_slug=org_slug)
    
    staff_members = organization.staff_members.all()
    context = {
        'organization': organization,
        'staff_members': staff_members,
        'step': 3,
        'total_steps': 5
    }
    return render(request, 'tracker/onboarding/staff.html', context)


@login_required
def onboarding_import(request, org_slug):
    """Step 4: Import Members (Optional)"""
    try:
        organization = Organization.objects.get(slug=org_slug)
    except Organization.DoesNotExist:
        messages.error(request, 'Organization not found.')
        return redirect('tracker:landing')
    
    if request.method == 'POST':
        # Check if user clicked "Skip & Go to Dashboard"
        if 'skip' in request.POST:
            return redirect('tracker:onboarding_subscription', org_slug=org_slug)
        
        # Handle file import
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            update_existing = form.cleaned_data['update_existing']
            default_pledge = form.cleaned_data['default_pledge']

            try:
                workbook = openpyxl.load_workbook(BytesIO(excel_file.read()))
                sheet = workbook.active

                created_count = 0
                updated_count = 0
                transaction_count = 0
                errors = []

                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not row[0]:
                            continue  # Skip rows without a name

                        name = str(row[0]).strip()
                        pledge = row[1]
                        paid = row[2] if len(row) > 2 else None
                        phone = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                        email = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                        course = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        year = str(row[6]).strip() if len(row) > 6 and row[6] else ''

                        # Pledge
                        try:
                            pledge = Decimal(str(pledge)) if pledge else default_pledge
                        except InvalidOperation:
                            pledge = default_pledge
                            errors.append(f"Row {idx}: Invalid pledge, using default for '{name}'")

                        # Paid
                        try:
                            paid = Decimal(str(paid)) if paid else Decimal('0.00')
                        except InvalidOperation:
                            paid = Decimal('0.00')
                            errors.append(f"Row {idx}: Invalid paid value for '{name}', using 0.00")

                        member_created = False
                        if update_existing:
                            member, member_created = Member.objects.update_or_create(
                                organization=organization,
                                name=name,
                                defaults={
                                    'pledge': pledge,
                                    'phone': phone or None,
                                    'email': email or None,
                                    'course': course or None,
                                    'year': year or None,
                                }
                            )
                        else:
                            member = Member.objects.filter(organization=organization, name=name).first()
                            if not member:
                                member = Member.objects.create(
                                    organization=organization,
                                    name=name,
                                    pledge=pledge,
                                    phone=phone or None,
                                    email=email or None,
                                    course=course or None,
                                    year=year or None,
                                )
                                member_created = True
                            else:
                                errors.append(f"Row {idx}: Member '{name}' already exists and update is off.")
                                continue

                        if member_created:
                            created_count += 1
                        else:
                            updated_count += 1

                        # Transaction
                        if paid > 0:
                            existing_txn = Transaction.objects.filter(
                                organization=organization,
                                member=member,
                                date=date.today(),
                                added_by=request.user,
                                note__icontains="Imported via Excel"
                            ).first()

                            if existing_txn:
                                # Update the amount if different
                                if existing_txn.amount != paid:
                                    existing_txn.amount = paid
                                    existing_txn.note = f"Updated via Excel on {date.today().isoformat()}"
                                    existing_txn.save()
                                    transaction_count += 1
                            else:
                                # Create new transaction
                                Transaction.objects.create(
                                    organization=organization,
                                    member=member,
                                    amount=paid,
                                    date=date.today(),
                                    added_by=request.user,
                                    note=f"Imported via Excel on {date.today().isoformat()}"
                                )
                                transaction_count += 1

                    except Exception as e:
                        errors.append(f"Row {idx}: Error processing member '{row[0]}' - {str(e)}")

                # Show success messages
                if created_count:
                    messages.success(request, f"✅ Created {created_count} new members.")
                if updated_count:
                    messages.info(request, f"🔄 Updated {updated_count} existing members.")
                if transaction_count:
                    messages.success(request, f"💰 Recorded {transaction_count} payments.")

                # Show errors if any
                if errors:
                    for err in errors[:5]:
                        messages.warning(request, f"⚠️ {err}")
                    if len(errors) > 5:
                        messages.warning(request, f"...and {len(errors) - 5} more issues.")

                # Redirect to subscription after successful import
                return redirect('tracker:onboarding_subscription', org_slug=org_slug)

            except Exception as e:
                messages.error(request, f"📄 Excel reading error: {str(e)}")

        else:
            messages.error(request, "⚠️ Invalid form submission.")
    else:
        form = ExcelImportForm()
    
    context = {
        'organization': organization,
        'step': 4,
        'total_steps': 5,
        'form': form
    }
    return render(request, 'tracker/onboarding/import.html', context)


@login_required
def onboarding_subscription(request, org_slug):
    """Step 5: Subscription & Payment (manual)"""
    try:
        organization = Organization.objects.get(slug=org_slug)
    except Organization.DoesNotExist:
        messages.error(request, 'Organization not found.')
        return redirect('tracker:landing')

    BASE_PRICE = Decimal('19800.00')
    category = organization.category
    category_discount = 50 if category == 'religious' else 35

    # Check if there's a pending payment request (only truly pending)
    pending_request = PaymentRequest.objects.filter(
        organization=organization,
        submitted_by=request.user,
        status='pending'
    ).order_by('-created_at').first()
    
    # Get the most recent declined request (if any) for prominent display
    recent_declined = PaymentRequest.objects.filter(
        organization=organization,
        submitted_by=request.user,
        status='declined'
    ).order_by('-created_at').first() if not pending_request else None
    
    # Get all payment requests for this organization (for history display)
    all_requests = PaymentRequest.objects.filter(
        organization=organization,
        submitted_by=request.user
    ).order_by('-created_at')[:10]  # Show last 10 requests
    
    # Check if there are any declined requests (for highlighting)
    has_declined = PaymentRequest.objects.filter(
        organization=organization,
        submitted_by=request.user,
        status='declined'
    ).exists()

    # Check if this organization was new when they submitted their request
    # This is determined by checking if there were NO payment requests before the current pending one
    # OR by checking session flag set during submission
    was_new_on_submit = False
    if pending_request:
        # Count how many requests existed BEFORE this pending one
        requests_before = PaymentRequest.objects.filter(
            organization=organization,
            created_at__lt=pending_request.created_at
        ).count()
        was_new_on_submit = (requests_before == 0)
    
    # Also check session flag (set during POST submission)
    if 'was_new_org_on_submit' in request.session:
        was_new_on_submit = request.session.pop('was_new_org_on_submit', False)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'trial':
            PaymentRequest.objects.create(
                organization=organization,
                submitted_by=request.user,
                months=0,
                is_trial=True,
                amount_tzs=Decimal('0.00'),
                discount_percent=0,
                category_snapshot=category,
                reference_note=None,
            )
            # Mark trial start and status
            if not organization.trial_started_at:
                organization.trial_started_at = timezone.now()
                organization.subscription_status = 'FREE_TRIAL'
                organization.save(update_fields=['trial_started_at', 'subscription_status'])
            messages.success(request, 'Free trial activated for 7 days. Enjoy!')
            return redirect('tracker:dashboard', org_slug=org_slug)

        # Paid subscription
        months = int(request.POST.get('months') or '1')
        months = max(1, min(12, months))
        first_month_discount = Decimal(str(category_discount))
        # Discount applies to first month only
        first_month_price = BASE_PRICE * (Decimal('100') - first_month_discount) / Decimal('100')
        total_amount = first_month_price + (BASE_PRICE * (months - 1))

        # Collect billing fields
        reference_note = (request.POST.get('payment_reference') or request.POST.get('reference_note') or '').strip() or None
        payment_method = request.POST.get('payment_method') or None
        amount_sent_raw = request.POST.get('amount_sent') or None
        amount_sent = None
        try:
            if amount_sent_raw is not None and amount_sent_raw != '':
                amount_sent = Decimal(str(amount_sent_raw))
        except Exception:
            amount_sent = None

        # Check if this is a new organization (no payment requests ever existed BEFORE creating this one)
        was_new_org = not PaymentRequest.objects.filter(organization=organization).exists()
        
        # Create payment request
        payment_request = PaymentRequest.objects.create(
            organization=organization,
            submitted_by=request.user,
            months=months,
            is_trial=False,
            amount_tzs=total_amount.quantize(Decimal('1.00')),
            discount_percent=int(category_discount),
            category_snapshot=category,
            reference_note=reference_note,
            payment_method=payment_method,
            amount_sent=amount_sent,
        )
        
        # For NEW organizations (no previous payment requests): activate free trial while payment is being processed
        # This allows them to access the dashboard immediately during onboarding
        if was_new_org and not organization.trial_started_at:
            organization.trial_started_at = timezone.now()
            organization.subscription_status = 'FREE_TRIAL'
            organization.save(update_fields=['trial_started_at', 'subscription_status'])
            messages.success(request, 'Your subscription request has been submitted! We\'ve activated a free 7-day trial while we process your payment. You\'ll get full access once approved.')
            # Auto-redirect new users to dashboard after activating free trial
            return redirect('tracker:dashboard', org_slug=org_slug)
        else:
            messages.success(request, 'Subscription request submitted successfully! Admin will confirm after payment verification.')
            # Auto-redirect to dashboard after submitting payment request
            return redirect('tracker:dashboard', org_slug=org_slug)

    context = {
        'organization': organization,
        'step': 5,
        'total_steps': 5,
        'base_price': 19800,
        'category_discount': category_discount,
        'pay_number': '68256127',
        'pending_request': pending_request,
        'recent_declined': recent_declined,
        'all_requests': all_requests,
        'has_declined': has_declined,
        'was_new_on_submit': was_new_on_submit,
        'whatsapp_number': '+255614021404',
        'support_email': 'kodinsoftwares@gmail.com',
    }
    return render(request, 'tracker/onboarding/subscription.html', context)

@login_required
def subscription_renewal(request, org_slug):
    """Subscription renewal page for expired subscriptions (no free trial option)"""
    try:
        organization = Organization.objects.get(slug=org_slug)
    except Organization.DoesNotExist:
        messages.error(request, 'Organization not found.')
        return redirect('tracker:landing')

    BASE_PRICE = Decimal('19800.00')
    category = organization.category
    category_discount = 50 if category == 'religious' else 35

    # Check if organization is already subscribed and active
    now = timezone.now()
    is_subscribed = (
        organization.subscription_status == 'SUBSCRIBED' and 
        organization.subscription_expires_at and 
        organization.subscription_expires_at > now
    )
    
    # If already subscribed, check for recently approved request and redirect to dashboard
    if is_subscribed:
        # Check if there's a recently approved request (within last 5 minutes)
        recent_approved = PaymentRequest.objects.filter(
            organization=organization,
            submitted_by=request.user,
            status='approved',
            is_trial=False,
            updated_at__gte=now - timezone.timedelta(minutes=5)
        ).order_by('-updated_at').first()
        
        if recent_approved:
            # Recently approved - show success message
            days_left = (organization.subscription_expires_at - now).days
            messages.success(
                request, 
                f'Your payment request has been approved! Your subscription is active until {organization.subscription_expires_at.strftime("%B %d, %Y")} ({days_left} days remaining).'
            )
        else:
            # Already subscribed but not recently approved - show info message
            days_left = (organization.subscription_expires_at - now).days
            messages.info(
                request, 
                f'Your subscription is active until {organization.subscription_expires_at.strftime("%B %d, %Y")} ({days_left} days remaining).'
            )
        
        # Redirect to dashboard
        return redirect('tracker:dashboard', org_slug=org_slug)

    # Check if there's a pending payment request
    pending_request = PaymentRequest.objects.filter(
        organization=organization,
        submitted_by=request.user,
        status='pending'
    ).order_by('-created_at').first()
    
    # Get the most recent declined request (if any) for prominent display
    recent_declined = PaymentRequest.objects.filter(
        organization=organization,
        submitted_by=request.user,
        status='declined'
    ).order_by('-created_at').first() if not pending_request else None
    
    # Get all payment requests for this organization (for history display)
    all_requests = PaymentRequest.objects.filter(
        organization=organization,
        submitted_by=request.user
    ).order_by('-created_at')[:10]

    if request.method == 'POST':
        # Only paid subscription (no free trial option)
        months = int(request.POST.get('months') or '1')
        months = max(1, min(12, months))
        first_month_discount = Decimal(str(category_discount))
        first_month_price = BASE_PRICE * (Decimal('100') - first_month_discount) / Decimal('100')
        total_amount = first_month_price + (BASE_PRICE * (months - 1))

        # Collect billing fields
        reference_note = (request.POST.get('payment_reference') or request.POST.get('reference_note') or '').strip() or None
        payment_method = request.POST.get('payment_method') or None
        amount_sent_raw = request.POST.get('amount_sent') or None
        amount_sent = None
        try:
            if amount_sent_raw is not None and amount_sent_raw != '':
                amount_sent = Decimal(str(amount_sent_raw))
        except Exception:
            amount_sent = None

        # Create payment request
        payment_request = PaymentRequest.objects.create(
            organization=organization,
            submitted_by=request.user,
            months=months,
            is_trial=False,
            amount_tzs=total_amount.quantize(Decimal('1.00')),
            discount_percent=int(category_discount),
            category_snapshot=category,
            reference_note=reference_note,
            payment_method=payment_method,
            amount_sent=amount_sent,
        )
        
        messages.success(request, 'Subscription request submitted successfully! Admin will confirm after payment verification.')
        return redirect('tracker:dashboard', org_slug=org_slug)

    context = {
        'organization': organization,
        'base_price': 19800,
        'category_discount': category_discount,
        'pay_number': '68256127',
        'pending_request': pending_request,
        'recent_declined': recent_declined,
        'all_requests': all_requests,
        'whatsapp_number': '+255614021404',
        'support_email': 'kodinsoftwares@gmail.com',
        'is_renewal': True,  # Flag to indicate this is renewal page (not onboarding)
    }
    return render(request, 'tracker/subscription.html', context)

@login_required
def staff_onboarding(request, org_slug):
    """Staff onboarding - collect email and full name for staff created by admin"""
    try:
        organization = Organization.objects.get(slug=org_slug)
    except Organization.DoesNotExist:
        messages.error(request, 'Organization not found.')
        return redirect('tracker:landing')

    # Check if user is staff and needs onboarding
    try:
        org_user = OrganizationUser.objects.get(user=request.user, organization=organization)
        if org_user.role not in ['staff', 'admin']:
            messages.error(request, 'Access denied.')
            return redirect('tracker:dashboard', org_slug=org_slug)
    except OrganizationUser.DoesNotExist:
        messages.error(request, 'You are not a member of this organization.')
        return redirect('tracker:landing')

    # Check if onboarding is already completed
    if request.user.userprofile.onboarding_completed:
        return redirect('tracker:dashboard', org_slug=org_slug)

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()

        errors = []

        # Validate email
        if not email:
            errors.append('Email is required.')
        elif User.objects.filter(email=email).exclude(id=request.user.id).exists():
            errors.append('This email is already in use by another account.')

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            # Update user profile
            request.user.email = email
            request.user.userprofile.phone = phone
            request.user.userprofile.onboarding_completed = True
            request.user.save()
            request.user.userprofile.save()

            messages.success(request, 'Profile completed successfully! Welcome to the team.')
            return redirect('tracker:dashboard', org_slug=org_slug)

    context = {
        'organization': organization,
        'user': request.user,
    }
    return render(request, 'tracker/staff_onboarding.html', context)


# ============================================================================
# ERROR HANDLERS
# ============================================================================

def custom_404(request, exception=None):
    """Custom 404 error page handler"""
    return render(request, '404.html', {'exception': exception}, status=404)


def custom_500(request):
    """Custom 500 error page handler"""
    return render(request, '500.html', status=500)


@login_required
def health_check(request):
    """Health check endpoint for monitoring"""
    from django.db import connection
    from django.core.cache import cache

    health_status = {
        'status': 'healthy',
        'timestamp': timezone.now().isoformat(),
        'checks': {}
    }

    # Database check
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        health_status['checks']['database'] = 'healthy'
    except Exception as e:
        health_status['status'] = 'unhealthy'
        health_status['checks']['database'] = f'unhealthy: {str(e)}'

    # Cache check (if using cache)
    try:
        cache.set('health_check', 'ok', 10)
        cache_value = cache.get('health_check')
        if cache_value == 'ok':
            health_status['checks']['cache'] = 'healthy'
        else:
            health_status['checks']['cache'] = 'unhealthy: cache not working'
    except:
        health_status['checks']['cache'] = 'disabled'

    # Application check
    health_status['checks']['application'] = 'healthy'

    status_code = 200 if health_status['status'] == 'healthy' else 503

    return JsonResponse(health_status, status=status_code)


@login_required
def export_admin_log_excel(request, org_slug=None):
    """Export admin log transactions to Excel"""
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect('tracker:dashboard')

    # Get filtered transactions
    transactions = Transaction.objects.select_related('member', 'added_by').filter(member__organization=tenant)

    # Apply filters
    boss_filter = request.GET.get('boss', '')
    date_filter = request.GET.get('date', '')
    amount_filter = request.GET.get('amount', '')

    if boss_filter:
        transactions = transactions.filter(added_by__username__icontains=boss_filter)
    if date_filter:
        transactions = transactions.filter(date=date_filter)
    if amount_filter:
        try:
            amount = Decimal(amount_filter)
            transactions = transactions.filter(amount=amount)
        except:
            pass

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Admin Log - Transactions"

    # Header
    headers = ['Date', 'Member Name', 'Amount (TZS)', 'Added By', 'Note']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    # Data
    for row, transaction in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=transaction.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=transaction.member.name)
        ws.cell(row=row, column=3, value=float(transaction.amount))
        ws.cell(row=row, column=4, value=transaction.added_by.username)
        ws.cell(row=row, column=5, value=transaction.note or '')

    # Auto-adjust column widths
    for col in range(1, 6):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(transactions) + 2):
            cell_value = str(ws[f'{column_letter}{row}'].value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="admin_log_{tenant.slug}_{date.today().isoformat()}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_admin_log_pdf(request, org_slug=None):
    """Export admin log transactions to PDF"""
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect('tracker:dashboard')

    # Get filtered transactions
    transactions = Transaction.objects.select_related('member', 'added_by').filter(member__organization=tenant)

    # Apply filters
    boss_filter = request.GET.get('boss', '')
    date_filter = request.GET.get('date', '')
    amount_filter = request.GET.get('amount', '')

    if boss_filter:
        transactions = transactions.filter(added_by__username__icontains=boss_filter)
    if date_filter:
        transactions = transactions.filter(date=date_filter)
    if amount_filter:
        try:
            amount = Decimal(amount_filter)
            transactions = transactions.filter(amount=amount)
        except:
            pass

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Get styles
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph(f"Admin Log - Transaction History<br/>{tenant.name}", title_style))
    elements.append(Spacer(1, 12))

    # Table data
    data = [['Date', 'Member', 'Amount', 'Added By', 'Note']]
    for transaction in transactions:
        data.append([
            transaction.date.strftime('%Y-%m-%d'),
            transaction.member.name,
            f"TZS {transaction.amount:,.0f}",
            transaction.added_by.username,
            transaction.note or ''
        ])

    # Table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7492b9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # Create table
    table = Table(data, colWidths=[80, 100, 80, 80, 150])
    table.setStyle(table_style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="admin_log_{tenant.slug}_{date.today().isoformat()}.pdf"'
    return response


@login_required
def export_member_edit_log_excel(request, org_slug=None):
    """Export member edit logs to Excel"""
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect('tracker:dashboard')

    # Check if user is owner
    is_owner = is_org_owner(request.user, tenant)
    if not is_owner:
        messages.error(request, 'Access denied. Owner privileges required.')
        return redirect('tracker:admin_log', org_slug=org_slug)

    # Get filtered member edit logs
    member_edit_logs = MemberEditLog.objects.select_related('member', 'edited_by').filter(
        organization=tenant
    ).order_by('-created_at')

    # Apply filters
    edit_member_filter = request.GET.get('edit_member', '')
    edit_field_filter = request.GET.get('edit_field', '')

    if edit_member_filter:
        member_edit_logs = member_edit_logs.filter(member__name__icontains=edit_member_filter)
    if edit_field_filter:
        member_edit_logs = member_edit_logs.filter(field_changed=edit_field_filter)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Member Edit Log"

    # Header
    headers = ['Date & Time', 'Member', 'Field Changed', 'Before', 'After', 'Done By']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    # Data
    for row, log in enumerate(member_edit_logs, 2):
        ws.cell(row=row, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=log.member.name)
        ws.cell(row=row, column=3, value=log.field_changed.title())
        ws.cell(row=row, column=4, value=log.before_value or '')
        ws.cell(row=row, column=5, value=log.after_value or '')
        ws.cell(row=row, column=6, value=log.edited_by.username)

    # Auto-adjust column widths
    for col in range(1, 7):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(member_edit_logs) + 2):
            cell_value = str(ws[f'{column_letter}{row}'].value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="member_edit_log_{tenant.slug}_{date.today().isoformat()}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_member_edit_log_pdf(request, org_slug=None):
    """Export member edit logs to PDF"""
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        messages.error(request, 'Organization not found')
        return redirect('tracker:dashboard')

    # Check if user is owner
    is_owner = is_org_owner(request.user, tenant)
    if not is_owner:
        messages.error(request, 'Access denied. Owner privileges required.')
        return redirect('tracker:admin_log', org_slug=org_slug)

    # Get filtered member edit logs
    member_edit_logs = MemberEditLog.objects.select_related('member', 'edited_by').filter(
        organization=tenant
    ).order_by('-created_at')

    # Apply filters
    edit_member_filter = request.GET.get('edit_member', '')
    edit_field_filter = request.GET.get('edit_field', '')

    if edit_member_filter:
        member_edit_logs = member_edit_logs.filter(member__name__icontains=edit_member_filter)
    if edit_field_filter:
        member_edit_logs = member_edit_logs.filter(field_changed=edit_field_filter)

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Get styles
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph(f"Member Edit Log<br/>{tenant.name}", title_style))
    elements.append(Spacer(1, 12))

    # Table data
    data = [['Date & Time', 'Member', 'Field Changed', 'Before', 'After', 'Done By']]
    for log in member_edit_logs:
        data.append([
            log.created_at.strftime('%Y-%m-%d %H:%M'),
            log.member.name,
            log.field_changed.title(),
            log.before_value or '-',
            log.after_value or '-',
            log.edited_by.username
        ])

    # Table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7492b9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # Create table
    table = Table(data, colWidths=[70, 80, 70, 60, 60, 60])
    table.setStyle(table_style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="member_edit_log_{tenant.slug}_{date.today().isoformat()}.pdf"'
    return response
