"""API views for Bossin Finance mobile application."""

from decimal import Decimal, InvalidOperation
from datetime import date
from io import BytesIO

from django.contrib.auth.models import User
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import generics, status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenRefreshView

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from tracker.models import (
    Member,
    Transaction,
    Organization,
    OrganizationUser,
    MemberEditLog,
    PaymentRequest,
    SystemSettings,
)
from tracker.api.mixins import TenantMixin, APIResponseMixin
from tracker.api.permissions import (
    IsOrgMember,
    IsOrgStaff,
    IsOrgAdmin,
    IsOrgOwner,
    CanRecordTransactions,
    ReadOnlyForViewer,
    SubscriptionActive,
)
from tracker.api.serializers import (
    UserSerializer,
    OrganizationSerializer,
    OrganizationUpdateSerializer,
    OrganizationThemeSerializer,
    OrganizationThemeUpdateSerializer,
    OrganizationMembershipSerializer,
    MemberSerializer,
    TransactionSerializer,
    RecordPaymentSerializer,
    BulkPaymentSerializer,
    MemberEditLogSerializer,
    StaffSerializer,
    AddStaffSerializer,
    UpdateStaffSerializer,
    PaymentRequestSerializer,
    CreatePaymentRequestSerializer,
    LoginSerializer,
    RegisterSerializer,
    StaffOnboardingSerializer,
)
from tracker.api.utils import (
    get_dashboard_stats,
    filter_members_queryset,
    get_subscription_pricing,
    calculate_subscription_amount,
    check_subscription_active,
    success_response,
    error_response,
)


def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }


# =============================================================================
# AUTH ENDPOINTS
# =============================================================================

class LoginAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if not serializer.is_valid():
            return self._error_response(serializer.errors)

        user = serializer.validated_data['user']
        default_org = serializer.validated_data['default_org']
        tokens = get_tokens_for_user(user)

        memberships = OrganizationUser.objects.filter(
            user=user, is_active=True,
        ).select_related('organization', 'organization__theme')

        return self._success_response({
            'tokens': tokens,
            'user': UserSerializer(user).data,
            'default_org': OrganizationSerializer(
                default_org, context={'request': request},
            ).data,
            'organizations': OrganizationSerializer(
                [m.organization for m in memberships],
                many=True,
                context={'request': request},
            ).data,
        })

    def _success_response(self, data, status_code=200):
        from tracker.api.utils import success_response
        return success_response(data, status=status_code)

    def _error_response(self, errors, status_code=400):
        from tracker.api.utils import error_response
        return error_response(errors, status=status_code)


class RegisterAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if not serializer.is_valid():
            return error_response(serializer.errors)

        result = serializer.save()
        user = result['user']
        organization = result['organization']
        tokens = get_tokens_for_user(user)

        return success_response({
            'tokens': tokens,
            'user': UserSerializer(user).data,
            'organization': OrganizationSerializer(
                organization, context={'request': request},
            ).data,
        }, status=status.HTTP_201_CREATED)


class MeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        memberships = OrganizationUser.objects.filter(
            user=request.user, is_active=True,
        ).select_related('organization', 'organization__theme')

        return success_response({
            'user': UserSerializer(request.user).data,
            'organizations': OrganizationSerializer(
                [m.organization for m in memberships],
                many=True,
                context={'request': request},
            ).data,
            'memberships': [
                {
                    'organization_slug': m.organization.slug,
                    'organization_name': m.organization.name,
                    'role': m.role,
                }
                for m in memberships
            ],
        })


class StaffOnboardingAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgMember]

    def post(self, request, org_slug):
        if request.user.userprofile.onboarding_completed:
            return self.api_error('Onboarding already completed.')

        serializer = StaffOnboardingSerializer(
            data=request.data, context={'request': request},
        )
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        request.user.email = serializer.validated_data['email']
        request.user.userprofile.phone = serializer.validated_data.get('phone', '')
        request.user.userprofile.onboarding_completed = True
        request.user.userprofile.needs_onboarding = False
        request.user.save()
        request.user.userprofile.save()

        return self.api_success({
            'user': UserSerializer(request.user).data,
            'message': 'Profile completed successfully.',
        })


class RefreshTokenAPIView(TokenRefreshView):
    """JWT token refresh - wraps simplejwt with consistent response."""

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == 200:
            return success_response({'tokens': response.data})
        return error_response(response.data, status=response.status_code)


# =============================================================================
# ORGANIZATION ENDPOINTS
# =============================================================================

class OrganizationListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        memberships = OrganizationUser.objects.filter(
            user=request.user, is_active=True,
        ).select_related('organization', 'organization__theme')

        return success_response({
            'organizations': OrganizationSerializer(
                [m.organization for m in memberships],
                many=True,
                context={'request': request},
            ).data,
        })


class OrganizationDetailAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgMember]

    def get(self, request, org_slug):
        return self.api_success(
            OrganizationSerializer(
                request.tenant, context={'request': request},
            ).data,
        )

    def patch(self, request, org_slug):
        if not IsOrgAdmin().has_permission(request, self):
            return self.api_error('You do not have admin permissions.', status=403)

        serializer = OrganizationUpdateSerializer(
            request.tenant, data=request.data, partial=True,
        )
        if not serializer.is_valid():
            return self.api_error(serializer.errors)
        serializer.save()
        return self.api_success(
            OrganizationSerializer(
                request.tenant, context={'request': request},
            ).data,
        )


class OrganizationThemeAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgMember]

    def get(self, request, org_slug):
        return self.api_success(
            OrganizationThemeSerializer(
                request.tenant.theme, context={'request': request},
            ).data,
        )

    def patch(self, request, org_slug):
        if not IsOrgAdmin().has_permission(request, self):
            return self.api_error('You do not have admin permissions.', status=403)

        serializer = OrganizationThemeUpdateSerializer(
            request.tenant.theme, data=request.data, partial=True,
            context={'request': request},
        )
        if not serializer.is_valid():
            return self.api_error(serializer.errors)
        serializer.save()
        return self.api_success(
            OrganizationThemeSerializer(
                request.tenant.theme, context={'request': request},
            ).data,
        )


# =============================================================================
# DASHBOARD
# =============================================================================

class DashboardStatsAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgMember]

    def get(self, request, org_slug):
        search = request.query_params.get('search', '')
        status_filter = request.query_params.get('filter', '')

        members_qs = Member.objects.filter(
            organization=request.tenant, is_active=True,
        )
        members_qs = filter_members_queryset(members_qs, search, status_filter)
        stats = get_dashboard_stats(request.tenant, members_qs)

        return self.api_success(stats)


# =============================================================================
# MEMBERS
# =============================================================================

class MemberListCreateAPIView(TenantMixin, APIResponseMixin, generics.ListCreateAPIView):
    serializer_class = MemberSerializer
    permission_classes = [
        IsAuthenticated, IsOrgMember, ReadOnlyForViewer,
        SubscriptionActive,
    ]

    def get_queryset(self):
        qs = Member.objects.filter(
            organization=self.request.tenant, is_active=True,
        ).order_by('name')
        search = self.request.query_params.get('search', '')
        status_filter = self.request.query_params.get('filter', '')
        return filter_members_queryset(qs, search, status_filter)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['organization'] = self.request.tenant
        return context

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({
                'success': True,
                'data': serializer.data,
                'error': None,
            })

        serializer = self.get_serializer(queryset, many=True)
        return self.api_success(serializer.data)

    def create(self, request, *args, **kwargs):
        if not IsOrgStaff().has_permission(request, self):
            return self.api_error('You do not have staff permissions.', status=403)

        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        member = serializer.save(organization=request.tenant)
        return self.api_success(
            MemberSerializer(member).data,
            status=status.HTTP_201_CREATED,
        )


class MemberDetailAPIView(TenantMixin, APIResponseMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = MemberSerializer
    permission_classes = [
        IsAuthenticated, IsOrgMember, ReadOnlyForViewer,
        SubscriptionActive,
    ]
    lookup_url_kwarg = 'member_id'

    def get_queryset(self):
        return Member.objects.filter(organization=self.request.tenant)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['organization'] = self.request.tenant
        return context

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return self.api_success(MemberSerializer(instance).data)

    def update(self, request, *args, **kwargs):
        if not IsOrgStaff().has_permission(request, self):
            return self.api_error('You do not have staff permissions.', status=403)

        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        old_values = {
            field: str(getattr(instance, field, ''))
            for field in ['name', 'phone', 'email', 'course', 'year', 'pledge']
        }

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        member = serializer.save()

        for field in old_values:
            new_val = str(getattr(member, field, ''))
            if old_values[field] != new_val:
                MemberEditLog.objects.create(
                    organization=request.tenant,
                    member=member,
                    field_changed=field,
                    before_value=old_values[field],
                    after_value=new_val,
                    edited_by=request.user,
                )

        return self.api_success(MemberSerializer(member).data)

    def destroy(self, request, *args, **kwargs):
        if not IsOrgStaff().has_permission(request, self):
            return self.api_error('You do not have staff permissions.', status=403)

        instance = self.get_object()
        instance.is_active = False
        instance.save(update_fields=['is_active', 'updated_at'])
        return self.api_success({'message': 'Member deactivated successfully.'})


class MemberTransactionsAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgMember]

    def get(self, request, org_slug, member_id):
        member = get_object_or_404(Member, id=member_id, organization=request.tenant)
        transactions = Transaction.objects.filter(
            member=member, organization=request.tenant,
        ).select_related('added_by').order_by('-date', '-created_at')

        return self.api_success(
            TransactionSerializer(transactions, many=True).data,
        )

    def post(self, request, org_slug, member_id):
        if not CanRecordTransactions().has_permission(request, self):
            return self.api_error('You do not have permission to record transactions.', status=403)

        sub_check = self.check_subscription_or_respond()
        if sub_check:
            return sub_check

        member = get_object_or_404(Member, id=member_id, organization=request.tenant)
        serializer = RecordPaymentSerializer(data=request.data)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        payment_date = serializer.validated_data.get('date') or timezone.now().date()
        transaction = Transaction.objects.create(
            organization=request.tenant,
            member=member,
            amount=serializer.validated_data['amount'],
            date=payment_date,
            added_by=request.user,
            note=serializer.validated_data.get('note', ''),
        )
        member.refresh_from_db()

        return self.api_success({
            'transaction': TransactionSerializer(transaction).data,
            'member': MemberSerializer(member).data,
        }, status=status.HTTP_201_CREATED)


# =============================================================================
# TRANSACTIONS
# =============================================================================

class TransactionListAPIView(TenantMixin, APIResponseMixin, generics.ListAPIView):
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated, IsOrgAdmin]

    def get_queryset(self):
        qs = Transaction.objects.filter(
            organization=self.request.tenant,
        ).select_related('member', 'added_by').order_by('-date', '-created_at')

        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        member_id = self.request.query_params.get('member_id')
        added_by = self.request.query_params.get('added_by')

        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if member_id:
            qs = qs.filter(member_id=member_id)
        if added_by:
            qs = qs.filter(added_by_id=added_by)

        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({
                'success': True,
                'data': serializer.data,
                'error': None,
            })
        serializer = self.get_serializer(queryset, many=True)
        return self.api_success(serializer.data)


class TransactionDetailAPIView(TenantMixin, APIResponseMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated, IsOrgAdmin, SubscriptionActive]
    lookup_url_kwarg = 'transaction_id'

    def get_queryset(self):
        return Transaction.objects.filter(organization=self.request.tenant)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return self.api_success(TransactionSerializer(instance).data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)
        transaction = serializer.save()
        return self.api_success(TransactionSerializer(transaction).data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return self.api_success({'message': 'Transaction deleted successfully.'})


class BulkPaymentAPIView(TenantMixin, APIResponseMixin, APIView):
    """Record multiple daily collection payments at once."""

    permission_classes = [IsAuthenticated, CanRecordTransactions, SubscriptionActive]

    def post(self, request, org_slug):
        serializer = BulkPaymentSerializer(data=request.data)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        results = []
        errors = []
        today = timezone.now().date()

        for payment in serializer.validated_data['payments']:
            try:
                member = Member.objects.get(
                    id=payment['member_id'],
                    organization=request.tenant,
                    is_active=True,
                )
                transaction = Transaction.objects.create(
                    organization=request.tenant,
                    member=member,
                    amount=payment['payment_amount'],
                    date=today,
                    added_by=request.user,
                    note=payment.get('note', ''),
                )
                member.refresh_from_db()
                results.append({
                    'transaction_id': transaction.id,
                    'member': MemberSerializer(member).data,
                })
            except Member.DoesNotExist:
                errors.append({
                    'member_id': payment['member_id'],
                    'error': 'Member not found.',
                })
            except Exception as e:
                errors.append({
                    'member_id': payment['member_id'],
                    'error': str(e),
                })

        return self.api_success({
            'recorded': results,
            'errors': errors,
            'success_count': len(results),
            'error_count': len(errors),
        })


# =============================================================================
# STAFF MANAGEMENT
# =============================================================================

class StaffListCreateAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgAdmin]

    def get(self, request, org_slug):
        staff = OrganizationUser.objects.filter(
            organization=request.tenant,
        ).select_related('user', 'user__userprofile').order_by('role', 'user__username')

        return self.api_success(StaffSerializer(staff, many=True).data)

    def post(self, request, org_slug):
        serializer = AddStaffSerializer(data=request.data)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        data = serializer.validated_data
        if User.objects.filter(username=data['username']).exists():
            return self.api_error('Username already exists.')

        user = User.objects.create_user(
            username=data['username'],
            email=data.get('email', ''),
            password=data['password'],
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
        )
        user.userprofile.needs_onboarding = True
        user.userprofile.save()

        org_user = OrganizationUser.objects.create(
            user=user,
            organization=request.tenant,
            role=data['role'],
            is_active=True,
        )

        return self.api_success(
            StaffSerializer(org_user).data,
            status=status.HTTP_201_CREATED,
        )


class StaffDetailAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgAdmin]

    def patch(self, request, org_slug, staff_id):
        org_user = get_object_or_404(
            OrganizationUser, id=staff_id, organization=request.tenant,
        )
        if org_user.role == 'owner':
            return self.api_error('Cannot modify organization owner.', status=403)

        serializer = UpdateStaffSerializer(data=request.data)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        if 'role' in serializer.validated_data:
            org_user.role = serializer.validated_data['role']
        if 'is_active' in serializer.validated_data:
            org_user.is_active = serializer.validated_data['is_active']
        org_user.save()

        return self.api_success(StaffSerializer(org_user).data)

    def delete(self, request, org_slug, staff_id):
        org_user = get_object_or_404(
            OrganizationUser, id=staff_id, organization=request.tenant,
        )
        if org_user.role == 'owner':
            return self.api_error('Cannot remove organization owner.', status=403)

        org_user.is_active = False
        org_user.save()
        return self.api_success({'message': 'Staff member deactivated.'})


# =============================================================================
# AUDIT LOG
# =============================================================================

class MemberEditLogAPIView(TenantMixin, APIResponseMixin, generics.ListAPIView):
    serializer_class = MemberEditLogSerializer
    permission_classes = [IsAuthenticated, IsOrgOwner]

    def get_queryset(self):
        return MemberEditLog.objects.filter(
            organization=self.request.tenant,
        ).select_related('member', 'edited_by').order_by('-created_at')

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({
                'success': True,
                'data': serializer.data,
                'error': None,
            })
        serializer = self.get_serializer(queryset, many=True)
        return self.api_success(serializer.data)


# =============================================================================
# SUBSCRIPTION & BILLING
# =============================================================================

class SubscriptionStatusAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgMember]

    def get(self, request, org_slug):
        is_active, status_info = check_subscription_active(request.tenant)
        pricing = get_subscription_pricing(request.tenant)

        pending = PaymentRequest.objects.filter(
            organization=request.tenant,
            submitted_by=request.user,
            status='pending',
        ).order_by('-created_at').first()

        recent_requests = PaymentRequest.objects.filter(
            organization=request.tenant,
            submitted_by=request.user,
        ).order_by('-created_at')[:10]

        return self.api_success({
            'subscription': status_info,
            'pricing': pricing,
            'pending_request': (
                PaymentRequestSerializer(pending).data if pending else None
            ),
            'recent_requests': PaymentRequestSerializer(
                recent_requests, many=True,
            ).data,
        })


class PaymentRequestCreateAPIView(TenantMixin, APIResponseMixin, APIView):
    permission_classes = [IsAuthenticated, IsOrgAdmin]

    def post(self, request, org_slug):
        serializer = CreatePaymentRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return self.api_error(serializer.errors)

        months = serializer.validated_data['months']
        total_amount, discount = calculate_subscription_amount(
            months, request.tenant,
        )

        payment_request = PaymentRequest.objects.create(
            organization=request.tenant,
            submitted_by=request.user,
            months=months,
            is_trial=False,
            amount_tzs=total_amount,
            discount_percent=discount,
            category_snapshot=request.tenant.category,
            reference_note=serializer.validated_data.get('reference_note'),
            payment_method=serializer.validated_data.get('payment_method'),
            amount_sent=serializer.validated_data.get('amount_sent'),
        )

        return self.api_success(
            PaymentRequestSerializer(payment_request).data,
            status=status.HTTP_201_CREATED,
        )


# =============================================================================
# SYSTEM / HEALTH
# =============================================================================

class HealthCheckAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        from django.db import connection

        db_ok = True
        try:
            with connection.cursor() as cursor:
                cursor.execute('SELECT 1')
        except Exception:
            db_ok = False

        return success_response({
            'status': 'healthy' if db_ok else 'unhealthy',
            'database': 'connected' if db_ok else 'disconnected',
            'timestamp': timezone.now().isoformat(),
            'version': '1.0.0',
        })


class HelpInfoAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        settings = SystemSettings.get_settings()
        return success_response({
            'support_email': settings.support_email,
            'support_phone': settings.support_phone,
            'mpesa_number': settings.mpesa_number,
            'mpesa_account_name': settings.mpesa_account_name,
            'whatsapp_number': '+255614021404',
        })


# =============================================================================
# EXPORT/IMPORT ENDPOINTS
# =============================================================================

class ExportMembersExcelAPIView(TenantMixin, APIView):
    """Export members data to Excel format."""
    permission_classes = [IsAuthenticated, IsOrgMember]

    def get(self, request, org_slug):
        search = request.query_params.get('search', '')
        filter_status = request.query_params.get('filter', '')

        # Get members with filtering
        members_qs = Member.objects.filter(
            organization=request.tenant, is_active=True,
        ).order_by('name')
        members_qs = filter_members_queryset(members_qs, search, filter_status)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members Data"

        # Add organization header
        ws.merge_cells('A1:J1')
        org_header = ws['A1']
        org_header.value = f"{request.tenant.name} - Members Report"
        org_header.font = Font(bold=True, size=14, color="FFFFFF")
        org_header.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        org_header.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Add export date
        ws.merge_cells('A2:J2')
        date_cell = ws['A2']
        date_cell.value = f"Exported on {date.today().strftime('%B %d, %Y')}"
        date_cell.font = Font(italic=True, size=10, color="666666")
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 15

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = ['Name', 'Pledge (TZS)', 'Paid (TZS)', 'Remaining (TZS)', 'Phone', 'Email', 'Course', 'Year', 'Status', 'Created Date']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Add data
        for row_num, member in enumerate(members_qs, 5):
            ws.cell(row=row_num, column=1, value=member.name)
            ws.cell(row=row_num, column=2, value=float(member.pledge))
            ws.cell(row=row_num, column=3, value=float(member.paid_total))
            ws.cell(row=row_num, column=4, value=float(member.remaining))
            ws.cell(row=row_num, column=5, value=member.phone or '')
            ws.cell(row=row_num, column=6, value=member.email or '')
            ws.cell(row=row_num, column=7, value=member.course or '')
            ws.cell(row=row_num, column=8, value=member.year or '')
            ws.cell(row=row_num, column=9, value=member.status_display)
            ws.cell(row=row_num, column=10, value=member.created_at.strftime('%Y-%m-%d'))

            # Add borders to data cells
            for col_num in range(1, 11):
                ws.cell(row=row_num, column=col_num).border = border

        # Adjust column widths
        column_widths = [30, 15, 15, 15, 15, 25, 20, 10, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{request.tenant.slug}_members_{date.today().strftime("%Y%m%d")}.xlsx"'
        return response


class ExportTransactionsExcelAPIView(TenantMixin, APIView):
    """Export transactions data to Excel format."""
    permission_classes = [IsAuthenticated, IsOrgAdmin]

    def get(self, request, org_slug):
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        # Get transactions
        transactions_qs = Transaction.objects.filter(
            organization=request.tenant,
        ).select_related('member', 'added_by').order_by('-date', '-created_at')

        if date_from:
            transactions_qs = transactions_qs.filter(date__gte=date_from)
        if date_to:
            transactions_qs = transactions_qs.filter(date__lte=date_to)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions Data"

        # Add organization header
        ws.merge_cells('A1:H1')
        org_header = ws['A1']
        org_header.value = f"{request.tenant.name} - Transactions Report"
        org_header.font = Font(bold=True, size=14, color="FFFFFF")
        org_header.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        org_header.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Add export date
        ws.merge_cells('A2:H2')
        date_cell = ws['A2']
        date_cell.value = f"Exported on {date.today().strftime('%B %d, %Y')}"
        date_cell.font = Font(italic=True, size=10, color="666666")
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 15

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = ['Date', 'Member Name', 'Amount (TZS)', 'Note', 'Added By', 'Created Date']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Add data
        for row_num, txn in enumerate(transactions_qs, 5):
            ws.cell(row=row_num, column=1, value=txn.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=2, value=txn.member.name if txn.member else 'N/A')
            ws.cell(row=row_num, column=3, value=float(txn.amount))
            ws.cell(row=row_num, column=4, value=txn.note or '')
            ws.cell(row=row_num, column=5, value=txn.added_by.username if txn.added_by else 'N/A')
            ws.cell(row=row_num, column=6, value=txn.created_at.strftime('%Y-%m-%d %H:%M'))

            # Add borders to data cells
            for col_num in range(1, 7):
                ws.cell(row=row_num, column=col_num).border = border

        # Adjust column widths
        column_widths = [15, 30, 15, 30, 20, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{request.tenant.slug}_transactions_{date.today().strftime("%Y%m%d")}.xlsx"'
        return response


class ExportReportPDFAPIView(TenantMixin, APIView):
    """Export members report to PDF format."""
    permission_classes = [IsAuthenticated, IsOrgMember]

    def get(self, request, org_slug):
        search = request.query_params.get('search', '')
        filter_status = request.query_params.get('filter', '')

        # Get members with filtering
        members_qs = Member.objects.filter(
            organization=request.tenant, is_active=True,
        ).order_by('name')
        members_qs = filter_members_queryset(members_qs, search, filter_status)

        # Get stats
        stats = get_dashboard_stats(request.tenant, members_qs)

        # Create PDF with proper margins
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{request.tenant.slug}_report_{date.today().strftime("%Y%m%d")}.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=120,
            bottomMargin=60
        )
        elements = []
        styles = getSampleStyleSheet()

        # Custom title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#2c3e50')
        )

        # Filter style
        filter_style = ParagraphStyle(
            'FilterStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#7f8c8d')
        )

        # Dynamic title based on filters
        base_title = f"MEMBERS REPORT: {request.tenant.name.upper()}"

        # Add filter information to title
        filter_info = []
        if filter_status:
            filter_map = {
                'incomplete': 'INCOMPLETE PAYMENTS',
                'complete': 'COMPLETE PAYMENTS',
                'pledged': 'PLEDGED ABOVE TSh 70,000',
                'not_started': 'NOT STARTED',
                'exceeded': 'EXCEEDED PLEDGES'
            }
            filter_info.append(filter_map.get(filter_status, filter_status.upper()))

        if search:
            filter_info.append(f'SEARCH: "{search.upper()}"')

        if filter_info:
            title_text = f"{base_title}<br/><font size='14' color='#e74c3c'>({' - '.join(filter_info)})</font>"
        else:
            title_text = base_title

        # Title
        title = Paragraph(title_text, title_style)
        elements.append(title)

        # Summary section (without report date)
        summary_data = [
            ['SUMMARY', ''],
            ['Total Members:', f"{stats['member_count']:,}"],
            ['Total Pledged:', f"TSh {float(stats['total_pledged']):,.2f}"],
            ['Total Collected:', f"TSh {float(stats['total_collected']):,.2f}"],
            ['Target Amount:', f"TSh {float(stats['target_amount']):,.2f}"],
            ['% Collected:', f"{stats['progress_percentage']:.1f}%"],
        ]

        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Members table with proper styling
        if members_qs:
            # Table headers with index number
            headers = ['#', 'Name', 'Phone', 'Pledge', 'Paid', 'Exceed/Remain(-)', 'Status']

            # Prepare data
            table_data = [headers]

            for index, member in enumerate(members_qs, 1):
                pledge = member.pledge if member.pledge is not None else Decimal('70000.00')
                paid = member.paid_total if member.paid_total is not None else Decimal('0.00')
                balance = paid - pledge

                if paid == 0:
                    status = "Not Started"
                elif paid < pledge:
                    status = "Incomplete"
                elif paid == pledge:
                    status = "Complete"
                else:
                    status = "Exceeded"

                row = [
                    str(index),  # Index number
                    member.name[:20] + "..." if len(member.name) > 20 else member.name,
                    member.phone or '',
                    f"TSh {float(pledge):,.0f}",
                    f"TSh {float(paid):,.0f}",
                    f"TSh {float(balance):,.0f}",
                    status,
                ]
                table_data.append(row)

            # Create table with proper column widths
            members_table = Table(
                table_data,
                colWidths=[0.5*inch, 2*inch, 1.2*inch, 1*inch, 1*inch, 1.2*inch, 1.1*inch]
            )
            members_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(members_table)

        doc.build(elements)
        return response


class ImportMembersExcelAPIView(TenantMixin, APIView):
    """Import members from Excel file."""
    permission_classes = [IsAuthenticated, IsOrgStaff, SubscriptionActive]

    def post(self, request, org_slug):
        if 'excel_file' not in request.FILES:
            return error_response('No file uploaded', status=400)

        excel_file = request.FILES['excel_file']
        update_existing = request.data.get('update_existing', 'false').lower() == 'true'
        default_pledge = request.data.get('default_pledge', '70000')

        try:
            default_pledge = Decimal(str(default_pledge))
        except (InvalidOperation, ValueError):
            default_pledge = Decimal('70000.00')

        try:
            workbook = openpyxl.load_workbook(BytesIO(excel_file.read()))
            sheet = workbook.active

            created_count = 0
            updated_count = 0
            transaction_count = 0
            errors = []

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:
                        continue  # Skip rows without a name

                    name = str(row[0]).strip()
                    pledge = row[1]
                    paid = row[2] if len(row) > 2 else None
                    phone = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    email = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    course = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                    year = str(row[6]).strip() if len(row) > 6 and row[6] else ''

                    # Pledge
                    try:
                        pledge = Decimal(str(pledge)) if pledge else default_pledge
                    except InvalidOperation:
                        pledge = default_pledge
                        errors.append(f"Row {idx}: Invalid pledge, using default for '{name}'")

                    # Paid
                    try:
                        paid = Decimal(str(paid)) if paid else Decimal('0.00')
                    except InvalidOperation:
                        paid = Decimal('0.00')
                        errors.append(f"Row {idx}: Invalid paid value for '{name}', using 0.00")

                    member_created = False
                    if update_existing:
                        member, member_created = Member.objects.update_or_create(
                            organization=request.tenant,
                            name=name,
                            defaults={
                                'pledge': pledge,
                                'phone': phone or None,
                                'email': email or None,
                                'course': course or None,
                                'year': year or None,
                            }
                        )
                    else:
                        member = Member.objects.filter(organization=request.tenant, name=name).first()
                        if not member:
                            member = Member.objects.create(
                                organization=request.tenant,
                                name=name,
                                pledge=pledge,
                                phone=phone or None,
                                email=email or None,
                                course=course or None,
                                year=year or None,
                            )
                            member_created = True
                        else:
                            errors.append(f"Row {idx}: Member '{name}' already exists and update is off.")
                            continue

                    if member_created:
                        created_count += 1
                    else:
                        updated_count += 1

                    # Transaction
                    if paid > 0:
                        existing_txn = Transaction.objects.filter(
                            organization=request.tenant,
                            member=member,
                            date=date.today(),
                            added_by=request.user,
                            note__icontains="Imported via Excel"
                        ).first()

                        if existing_txn:
                            # Update the amount if different
                            if existing_txn.amount != paid:
                                existing_txn.amount = paid
                                existing_txn.note = f"Updated via Excel on {date.today().isoformat()}"
                                existing_txn.save()
                                transaction_count += 1
                        else:
                            # Create new transaction
                            Transaction.objects.create(
                                organization=request.tenant,
                                member=member,
                                amount=paid,
                                date=date.today(),
                                added_by=request.user,
                                note=f"Imported via Excel on {date.today().isoformat()}"
                            )
                            transaction_count += 1

                except Exception as e:
                    errors.append(f"Row {idx}: Error processing member '{row[0]}' - {str(e)}")

            return success_response({
                'created_count': created_count,
                'updated_count': updated_count,
                'transaction_count': transaction_count,
                'errors': errors[:10],  # Limit errors to prevent huge responses
                'total_errors': len(errors),
            })

        except Exception as e:
            return error_response(f'Excel processing error: {str(e)}', status=400)
